import html
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_from_directory
import sqlite3
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, date
import os
import csv
import io
# import mysql.connector
from datetime import timedelta, datetime
from flask import send_file , Response 
from functools import wraps
from werkzeug.utils import secure_filename
# === REPORT / EXPORT ===
from openpyxl import Workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from weasyprint import HTML, CSS
from uuid import uuid4
from reportlab.lib.pagesizes import A4, landscape
import smtplib
from email.message import EmailMessage
from uuid import uuid4
from datetime import datetime, timedelta
from werkzeug.security import generate_password_hash, check_password_hash


PAGE_SIZE = landscape(A4)
app = Flask(__name__)

app.secret_key = os.environ.get("SECRET_KEY")

# === Upload config untuk dokumen sokongan leave ===
LEAVE_UPLOAD_FOLDER = os.path.join("static", "uploads", "leave_docs")
os.makedirs(LEAVE_UPLOAD_FOLDER, exist_ok=True)

ALLOWED_LEAVE_EXTENSIONS = {"png", "jpg", "jpeg", "pdf"}

# === Upload config untuk profile photo ===
PROFILE_UPLOAD_FOLDER = os.path.join("static", "uploads", "profile_photos")
app.config["PROFILE_UPLOAD_FOLDER"] = PROFILE_UPLOAD_FOLDER
os.makedirs(PROFILE_UPLOAD_FOLDER, exist_ok=True)


# file types allowed for profile photos
ALLOWED_EXTENSIONS = {"jpg", "jpeg", "png", "gif"}

app.config["PROFILE_UPLOAD_FOLDER"] = PROFILE_UPLOAD_FOLDER

def allowed_leave_file(filename):
    return (
        "." in filename
        and filename.rsplit(".", 1)[1].lower() in ALLOWED_LEAVE_EXTENSIONS
    )

DB_PATH = os.path.join(os.path.dirname(__file__), "database.db")
# ---------------------- Position Hierarchy ----------------------
POSITION_HIERARCHY = {
    "Staff": "Supervisor",
    "Supervisor": "Manager",
    "Manager": "General Manager",
    "General Manager": "CEO",
    "CEO": None  # top of chain
}

def calculate_working_days(start_date, end_date):
    """Return number of working days between start & end, 
    excluding Sat/Sun & public holidays."""
    conn = get_db()
    c = conn.cursor()
    c.execute(
        ("SELECT date FROM holidays")
    )
    holiday_rows = c.fetchall()
    holidays = {h["date"] for h in holiday_rows}
    conn.close()

    start = datetime.strptime(start_date, "%Y-%m-%d").date()
    end = datetime.strptime(end_date, "%Y-%m-%d").date()

    count = 0
    current = start

    while current <= end:
        if current.weekday() < 5:  # 0=Mon ... 4=Fri
            if current.isoformat() not in holidays:
                count += 1
        current += timedelta(days=1)
    return count

def get_used_leave_days(user_id, year=None):
    conn = get_db()
    c = conn.cursor()

    sql = """
        SELECT COALESCE(SUM(total_days), 0)
        FROM leave_applications
        WHERE user_id = %s
          AND status = 'Approved'
          AND leave_type != 'MC'
    """
    params = [user_id]

    if year:
        sql += " AND EXTRACT(YEAR FROM start_date) = %s"
        params.append(str(year))

    c.execute((sql), params)
    used = c.fetchone()["total"]
    conn.close()

    return used

def get_next_position(position):
    """Return next higher position for approval or checking chain."""
    return POSITION_HIERARCHY.get(position, None)

def allowed_photo(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

import psycopg2
from psycopg2.extras import RealDictCursor
import os

# def get_db():
#     database_url = os.environ.get("DATABASE_URL")

#     if database_url:
#         return psycopg2.connect(
#             database_url,
#             sslmode="require",
#             cursor_factory=RealDictCursor
#         )

#     # fallback for local
#     conn = sqlite3.connect("database.db")
#     conn.row_factory = sqlite3.Row
#     return conn

def get_db():
    database_url = os.environ.get("DATABASE_URL")
    return psycopg2.connect(
        database_url,
        sslmode="require",
        cursor_factory=RealDictCursor
    )

# def is_postgres():
#     return os.environ.get("DATABASE_URL") is not None

# def adapt_query(query):
#     """
#     Convert SQLite '?' placeholders to PostgreSQL '%s'
#     if using PostgreSQL.
#     """
#     if is_postgres():
#         return query.replace("?", "%s")
#     return query

# def _add_column_if_missing(cur, table, name, coltype):
#     if is_postgres():
#         cur.execute("""SELECT column_name
#             FROM information_schema.columns
#             WHERE table_name=%s
#         """, (table,))
#         cols = [r[0] for r in cur.fetchall()]
#     else:
#         cur.execute(f"PRAGMA table_info({table})")
#         cols = [r["name"] for r in cur.fetchall()]

#     if name not in cols:
#         cur.execute(f"ALTER TABLE {table} ADD COLUMN {name} {coltype}")


def init_db():
    conn = get_db()
    c = conn.cursor()

    c.execute("""
    CREATE TABLE IF NOT EXISTS departments (
        id SERIAL PRIMARY KEY,
        name VARCHAR(255) UNIQUE NOT NULL
    )
    """)

    c.execute("""
    CREATE TABLE IF NOT EXISTS users (
        id SERIAL PRIMARY KEY,
        username VARCHAR(150) UNIQUE NOT NULL,
        full_name VARCHAR(255) NOT NULL,
        password_hash TEXT NOT NULL,
        role VARCHAR(20) NOT NULL CHECK(role IN ('admin','user')),
        created_at TIMESTAMP NOT NULL,
        entitlement INTEGER DEFAULT 0,
        department_id INTEGER REFERENCES departments(id),
        position VARCHAR(100),
        approver_role VARCHAR(100),
        ic_number VARCHAR(50),
        email VARCHAR(255),
        phone VARCHAR(50),
        address TEXT,
        enrollment_date DATE,
        availability VARCHAR(50) DEFAULT 'Available',
        profile_photo TEXT,
        reset_token TEXT,
        reset_token_expiry TIMESTAMP
    )
    """)

    c.execute("""
    CREATE TABLE IF NOT EXISTS leaves (
        id SERIAL PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        leave_type TEXT NOT NULL,
        start_date TEXT NOT NULL,
        end_date TEXT NOT NULL,
        reason TEXT,
        status TEXT NOT NULL DEFAULT 'Pending',
        created_at TEXT NOT NULL,
        next_approver TEXT,
        FOREIGN KEY (user_id) REFERENCES users(id)
    )
    """)
    c.execute("""
    CREATE TABLE IF NOT EXISTS holidays (
        id SERIAL PRIMARY KEY,
        name VARCHAR(255),
        date DATE
    )
    """)
    c.execute("""
    CREATE TABLE IF NOT EXISTS settings (
        key TEXT PRIMARY KEY,
        value TEXT
    )
    """)
    c.execute("""
    CREATE TABLE IF NOT EXISTS departments (
        id SERIAL PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE NOT NULL
    )
    """)
    
    c.execute("""
    CREATE TABLE IF NOT EXISTS leave_logs (
        id SERIAL PRIMARY KEY AUTOINCREMENT,
        leave_id INTEGER NOT NULL,
        action TEXT NOT NULL,
        performed_by INTEGER NOT NULL,
        timestamp TEXT NOT NULL,
        description TEXT,
        FOREIGN KEY (leave_id) REFERENCES leaves(id),
        FOREIGN KEY (performed_by) REFERENCES users(id)
    )
    """)
    c.execute("""
    CREATE TABLE IF NOT EXISTS leave_applications (
        id SERIAL PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        full_name TEXT,
        position TEXT,
        leave_type TEXT NOT NULL,
        start_date TEXT NOT NULL,
        end_date TEXT NOT NULL,
        total_days INTEGER,
        reason TEXT,
        status TEXT NOT NULL DEFAULT 'Pending Recommender',
        checker_name TEXT,
        approver_name TEXT,
        support_doc TEXT,
        contact_address TEXT,
        contact_phone TEXT,
        created_at TEXT NOT NULL,
        checked_at TEXT,
        approved_at TEXT,
        FOREIGN KEY (user_id) REFERENCES users(id)
    )
    """)
    c.execute("""
    CREATE TABLE IF NOT EXISTS mc_records (
        id SERIAL PRIMARY KEY,
        user_id INTEGER REFERENCES users(id),
        mc_number VARCHAR(100),
        start_date DATE,
        end_date DATE,
        pdf_path TEXT,
        uploaded_by INTEGER REFERENCES users(id),
        created_at TIMESTAMP NOT NULL
    )
    """)

    # # Migrations (safe if already applied)
    # _add_column_if_missing(c, "users", "entitlement", "INTEGER DEFAULT 0")
    # _add_column_if_missing(c, "users", "department_id", "INTEGER REFERENCES departments(id)")
    # _add_column_if_missing(c, "users", "position", "TEXT")
    # _add_column_if_missing(c, "users", "approver_role", "TEXT")
    
    conn.commit()

    conn = get_db()
    c = conn.cursor()

    # Safely add new columns if they don't exist yet (for users table)
    def add_col_if_not_exists(colname, coldef):
        try:
            c.execute((f"ALTER TABLE users ADD COLUMN {colname} {coldef}"))
        except:
            pass  # column already exists

    add_col_if_not_exists("ic_number", "TEXT")
    add_col_if_not_exists("email", "TEXT")
    add_col_if_not_exists("phone", "TEXT")
    add_col_if_not_exists("address", "TEXT")
    add_col_if_not_exists("enrollment_date", "TEXT")
    add_col_if_not_exists("availability", "TEXT DEFAULT 'Available'")
    add_col_if_not_exists("profile_photo", "TEXT")
    add_col_if_not_exists("reset_token", "TEXT")
    add_col_if_not_exists("reset_token_expiry", "TEXT")


    # 🌿 NEW: Safely add columns to the leaves table
    def add_leaves_col_if_not_exists(colname, coldef):
        try:
            c.execute((f"ALTER TABLE leaves ADD COLUMN {colname} {coldef}"))
        except:
            pass  # column already exists

    add_leaves_col_if_not_exists("contact_address", "TEXT")
    add_leaves_col_if_not_exists("contact_phone", "TEXT")
    add_leaves_col_if_not_exists("notes", "TEXT")
    # add_leaves_col_if_not_exists("next_approver", "TEXT")
    add_leaves_col_if_not_exists("checked_by_position", "TEXT")
    add_leaves_col_if_not_exists("checked_status", "TEXT DEFAULT 'Pending'")
    add_leaves_col_if_not_exists("next_approver_position", "TEXT")
    add_leaves_col_if_not_exists("next_approver_department", "TEXT")
    add_leaves_col_if_not_exists("checked_by_user_id", "INTEGER")
    add_leaves_col_if_not_exists("approved_by_user_id", "INTEGER")

    conn.commit()

    # Seed admin + sample user
    c.execute(
        ("SELECT 1 FROM users WHERE username=%s"), 
        ('admin',)
    )
    if not c.fetchone():
        c.execute(
            """
            INSERT INTO users (username, full_name, password_hash, role, created_at, entitlement)
            VALUES (%s,%s,%s,%s,%s,%s)
            """,
            ("admin", "Administrator", generate_password_hash("admin123"), "admin",
            datetime.utcnow(), 20)
        )

    c.execute(
        ("SELECT 1 FROM users WHERE username=%s"), 
        ('user',)
    )
    if not c.fetchone():
        c.execute(
            (
            "INSERT INTO users (username, full_name, password_hash, role, created_at, entitlement) VALUES (%s,%s,%s,%s,%s,%s)"),
            ("user", "UserDash", generate_password_hash("password"), "user",
            datetime.utcnow(), 14)
        )

    # Seed simple holidays (only if empty)
    c.execute(
        ("SELECT COUNT(*) AS cnt FROM holidays"))
    if c.fetchone()["cnt"] == 0:
        holidays = [
            ("New Year's Day", f"{date.today().year}-01-01"),
            ("Labour Day", f"{date.today().year}-05-01"),
            ("Malaysia Day", f"{date.today().year}-09-16"),
        ]
        c.executemany("INSERT INTO holidays (name, date) VALUES (%s,%s)", holidays)

    conn.commit()
    conn.close()

def build_reset_email(reset_link):
    return f"""
    <html>
    <body style="font-family: Arial, sans-serif; background:#f4f6f8; padding:20px;">
      <table width="100%" cellpadding="0" cellspacing="0">
        <tr>
          <td align="center">
            <table width="600" style="background:#ffffff; border-radius:10px; padding:30px; box-shadow:0 5px 15px rgba(0,0,0,0.1);">
              
              <tr>
                <td align="center">
                  <h2 style="color:#1d73e8;">Reset Your Password</h2>
                </td>
              </tr>

              <tr>
                <td style="color:#333; font-size:15px;">
                  <p>Hello,</p>
                  <p>
                    You requested to reset your password for your account.
                    Click the button below to create a new password.
                  </p>

                  <p style="text-align:center; margin:30px 0;">
                    <a href="{reset_link}"
                       style="background:#1d73e8; color:white; padding:12px 25px;
                              text-decoration:none; border-radius:6px;
                              font-weight:bold;">
                      Reset Password
                    </a>
                  </p>

                  <p>
                    This link will expire in <b>1 hour</b>.
                    If you did not request this, please ignore this email.
                  </p>

                  <hr style="border:none; border-top:1px solid #eee;">

                  <p style="font-size:12px; color:#777;">
                    J-Leave System<br>
                    This is an automated email. Please do not reply.
                  </p>
                </td>
              </tr>

            </table>
          </td>
        </tr>
      </table>
    </body>
    </html>
    """

import smtplib
from email.message import EmailMessage

def send_email_html(to_email, subject, html_content):
    EMAIL_ADDRESS = os.environ.get("EMAIL_ADDRESS")
    EMAIL_PASSWORD = os.environ.get("EMAIL_PASSWORD")

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = EMAIL_ADDRESS
    msg["To"] = to_email

    msg.set_content("Please use an email client that supports HTML.")
    msg.add_alternative(html_content, subtype="html")

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
        smtp.send_message(msg)

@app.before_request
def ensure_db():
    if not hasattr(app, "_db_initialized"):
        init_db()
        app._db_initialized = True
    
def auto_reset_mc_availability():
    today = date.today().isoformat()
    conn = get_db()
    c = conn.cursor()

    c.execute(
        ("""SELECT DISTINCT u.id
            FROM mc_records m
            JOIN users u ON u.id = m.user_id
            WHERE m.end_date IS NOT NULL
            AND m.end_date < %s
            AND u.availability = 'MC'
        """), 
        (today,)
    )

    users_to_reset = c.fetchall()

    for u in users_to_reset:
        c.execute(
            ("UPDATE users SET availability='Available' WHERE id=%s"),
            (u["id"],)
        )

    conn.commit()
    conn.close()

@app.before_request
def before_any_request():
    auto_reset_mc_availability()

# ---------------------- Auth ----------------------
@app.route("/")
def home():
    if "user_id" in session:
        return redirect(url_for("admin_dashboard" if session.get("role") == "admin" else "user_dashboard"))
    return redirect(url_for("login"))

@app.route("/login", methods=["GET", "POST"])
def login():

    if "login_attempts" not in session:
        session["login_attempts"] = 0

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")


        if session["login_attempts"] >= 3:
            flash("Too many failed attempts. Please use 'Forgot password / username'.", "warning")
            return redirect(url_for("login"))

        conn = get_db()
        c = conn.cursor()
        c.execute(
            ("SELECT * FROM users WHERE username=%s"), 
            (username,)
        )
        user = c.fetchone()
        conn.close()

        # User not found
        if not user:
            session["login_attempts"] += 1
            flash("Invalid username or password.", "danger")
            return redirect(url_for("login"))

        # User resigned (treated as not exist)
        if user["availability"] == "Resign":
            flash("Your account is no longer active. Please contact admin.", "danger")
            return redirect(url_for("login"))

        # Wrong password
        if not check_password_hash(user["password_hash"], password):
            session["login_attempts"] += 1
            flash("Invalid username or password.", "danger")
            return redirect(url_for("login"))

        # success
        session["login_attempts"] = 0
        session.update({
            "user_id": user["id"],
            "username": user["username"],
            "full_name": user["full_name"],
            "role": user["role"],
            "position": user["position"],
            "profile_photo": user["profile_photo"]
        })

        flash("Login success. Welcome back!", "success")

        if user["role"] == "admin":
            return redirect(url_for("admin_dashboard"))

        elif (user["position"] or "").upper() == "CEO":
            return redirect(url_for("ceo_dashboard"))

        else:
            return redirect(url_for("user_dashboard"))

    return render_template(
        "login.html",
        show_forgot=session.get("login_attempts", 0) >= 3
    )

@app.route("/logout")
def logout():
    session.clear()
    flash("You have been logged out.", "info")
    return redirect(url_for("login"))

# ---------------------- Decorators ----------------------
def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if session.get("role") != "admin":
            flash("Admin access only.", "warning")
            return redirect(url_for("login"))
        return fn(*args, **kwargs)
    return wrapper

@app.route("/admin/mc/update/<int:mc_id>", methods=["POST"])
@admin_required
def admin_update_mc(mc_id):
    mc_number = request.form.get("mc_number")
    start = request.form.get("start_date")
    end = request.form.get("end_date")

    conn = get_db()
    c = conn.cursor()
    c.execute(
        ("""UPDATE mc_records
            SET mc_number=%s, start_date=%s, end_date=%s
            WHERE id=%s
        """), 
        (mc_number, start, end, mc_id)
    )
    conn.commit()
    conn.close()

    flash("MC updated.", "success")
    return redirect(url_for("admin_dashboard"))

def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not session.get("user_id"):
            flash("Please login first.", "warning")
            return redirect(url_for("login"))
        return fn(*args, **kwargs)
    return wrapper

@app.route("/admin/leaves/<view>")
def admin_leave_modal(view):
    if view == "today":
        data = get_today_leaves() 
    elif view == "pending":
        data = get_pending_leaves()
    else:
        data = get_all_leaves()

    return render_template("admin/modal_table.html", data=data)


@app.route("/admin/dashboard")
@admin_required
def admin_dashboard():

    # ================= FILTERS FROM QUERY STRING =================
    date_from = request.args.get("date_from", "").strip()
    date_to   = request.args.get("date_to", "").strip()

    year_filter  = request.args.get("year", "all")
    month_filter = request.args.get("month", "all")
    dept_filter  = request.args.get("department", "all")

    # ================= MAIN DASHBOARD DATA =================
    data = get_dashboard_data(
        date_from=date_from or None,
        date_to=date_to or None
    )

    conn = get_db()
    cur = conn.cursor()

    # ================= BUILD WHERE CONDITIONS =================
    conditions = ["l.status = 'Approved'"]
    params = []

    if year_filter != "all":
        conditions.append("EXTRACT(YEAR FROM l.start_date) = %s")
        params.append(year_filter)

    if month_filter != "all":
        conditions.append("EXTRACT(MONTH FROM l.start_date) = %s")
        params.append(month_filter)

    if dept_filter != "all":
        conditions.append("d.name = %s")
        params.append(dept_filter)

    where_sql = " AND ".join(conditions)

    # ================= TREND GRAPH DATA =================
    cur.execute(
        (f"""SELECT 
            EXTRACT(YEAR FROM l.start_date) AS year,
            EXTRACT(MONTH FROM l.start_date) AS month,
            COALESCE(d.name,'Unknown') AS department,
            COUNT(*) AS total
        FROM leave_applications l
        JOIN users u ON u.id = l.user_id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE {where_sql}
        GROUP BY year, month, department
        ORDER BY year, month
    """), params
    )

    trend_raw = [dict(r) for r in cur.fetchall()]

    # ================= DROPDOWN VALUES =================
    cur.execute(
        ("SELECT DISTINCT EXTRACT(YEAR FROM start_date) AS year FROM leave_applications"))
    years = [r["year"] for r in cur.fetchall() if r["year"]]

    cur.execute(
        ("SELECT DISTINCT EXTRACT(MONTH FROM start_date) AS month FROM leave_applications"))
    months = [r["month"] for r in cur.fetchall() if r["month"]]

    cur.execute(
        ("SELECT name FROM departments"))
    departments = [r["name"] for r in cur.fetchall()]

    conn.close()

    # ================= PASS TO TEMPLATE =================
    data["trend_raw"] = trend_raw
    data["years"] = years
    data["months"] = months
    data["departments"] = departments

    data["date_from"] = date_from
    data["date_to"] = date_to
    data["selected_year"] = year_filter
    data["selected_month"] = month_filter
    data["selected_department"] = dept_filter

    return render_template("admin_dashboard.html", **data)

@app.route("/export/leave/pdf")
def export_pdf():
    mode = request.args.get("mode","weekly")
    data = get_leave_data(mode)
    stream = io.BytesIO()
    pdf = canvas.Canvas(stream, pagesize=A4)
    pdf.setFont("Helvetica-Bold",16); pdf.drawString(40,800,f"Leave Report ({mode.title()})")
    pdf.setFont("Helvetica",10); y=770
    pdf.drawString(40,y,"Name"); pdf.drawString(180,y,"Type")
    pdf.drawString(260,y,"Start"); pdf.drawString(340,y,"End"); pdf.drawString(420,y,"Status"); y-=20
    for r in data:
        pdf.drawString(40,y,r["full_name"])
        pdf.drawString(180,y,r["leave_type"])
        pdf.drawString(260,y,r["start_date"])
        pdf.drawString(340,y,r["end_date"])
        pdf.drawString(420,y,r["status"])
        y -= 20
    pdf.save(); stream.seek(0)
    return send_file(stream, as_attachment=True,
                     download_name=f"LeaveReport_{mode}.pdf")
    

def export_leave_excel(data, year):
    wb = Workbook()
    ws = wb.active
    ws.title = "Leave Report"

    headers = ["ID","Name","Leave Type"] + MONTHS + ["Total Used","Remaining"]
    ws.append(headers)

    for r in data:
        row = [
            r["user_id"],
            r["name"],
            r["leave_type"],
            *[r["months"][m] for m in MONTHS],
            r["total_used"],
            r["remaining"]
        ]
        ws.append(row)

    file = io.BytesIO()
    wb.save(file)
    file.seek(0)

    return send_file(file,
        as_attachment=True,
        download_name=f"Leave_Report_{year}.xlsx")


from flask import make_response, render_template
from weasyprint import HTML
import io

@app.route("/admin/leave-report/employee/<int:user_id>/download")
@admin_required
def download_employee_leave_report(user_id):

    year = request.args.get("year")
    if not year:
        year = str(datetime.now().year)

    conn = get_db()
    c = conn.cursor()

    # ================= EMPLOYEE =================
    c.execute(
        ("""SELECT u.full_name, u.position, u.entitlement, d.name AS department
            FROM users u
            LEFT JOIN departments d ON u.department_id = d.id
            WHERE u.id = %s
        """), 
        (user_id,)
    )
    emp = c.fetchone()

    # ================= APPROVED LEAVES =================
    c.execute(
        ("""SELECT leave_type, start_date, end_date, total_days
            FROM leave_applications
            WHERE user_id = %s
            AND status='Approved'
            AND EXTRACT(YEAR FROM start_date) = %s
            ORDER BY start_date
        """), 
        (user_id, year)
    )
    approved_leaves = c.fetchall()

    # ================= MC RECORDS =================
    c.execute(
        ("""SELECT mc_number, start_date, end_date
            FROM mc_records
            WHERE user_id = %s
            AND EXTRACT(YEAR FROM start_date) = %s
            ORDER BY start_date
        """), 
        (user_id, year)
    )
    mc_records = c.fetchall()

    conn.close()

    # ================= MONTHS =================
    months = ["JAN","FEB","MAR","APR","MAY","JUN",
              "JUL","AUG","SEP","OCT","NOV","DEC"]

    month_map = {
        "JAN":"01","FEB":"02","MAR":"03","APR":"04",
        "MAY":"05","JUN":"06","JUL":"07","AUG":"08",
        "SEP":"09","OCT":"10","NOV":"11","DEC":"12"
    }

    # ================= MONTHLY =================
    monthly = {m:{} for m in months}

    for l in approved_leaves:
        if not l["start_date"] or not l["end_date"]:
            continue
        if l["leave_type"] == "MC":
            continue

        start = datetime.strptime(l["start_date"], "%Y-%m-%d").date()
        end   = datetime.strptime(l["end_date"], "%Y-%m-%d").date()

        d = start
        while d <= end:
            if d.year == int(year) and d.weekday() < 5:
                m = months[d.month-1]
                monthly[m][l["leave_type"]] = monthly[m].get(l["leave_type"],0)+1
            d += timedelta(days=1)

    entitled = emp["entitlement"] or 0
    used = sum(sum(v.values()) for v in monthly.values())
    balance = max(0, entitled-used)

    summary = {
        "entitled": entitled,
        "used": used,
        "balance": balance
    }

    leave_types = sorted(
        {l["leave_type"] for l in approved_leaves if l["leave_type"] != "MC"}
    )

    # Render HTML
    html = render_template(
        "reports/individual_leave_report.html",
        employee={
            "name": emp["full_name"],
            "position": emp["position"]
        },
        department=emp["department"],
        year=year,
        printed_date=datetime.now().strftime("%d %b %Y"),
        current_year=datetime.now().year,
        summary=summary,
        months=months,
        month_map=month_map,
        monthly=monthly,
        leave_types=leave_types,
        approved_leaves=approved_leaves,
        mc_records=mc_records,
        pdf_mode=True   # optional flag if needed
    )

    pdf = HTML(string=html, base_url=request.root_url).write_pdf()

    response = make_response(pdf)
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = f"attachment; filename=Leave_Report_{emp['full_name']}_{year}.pdf"

    return response


from datetime import timedelta

def daterange(start, end):
    for n in range((end - start).days + 1):
        yield start + timedelta(days=n)

@app.route("/leave-report/individual/<int:user_id>")
def individual_leave_preview(user_id):
    year = request.args.get("year", datetime.now().year)

    report = build_individual_leave_report(user_id, year)

    return render_template(
        "reports/individual_leave_preview.html",
        employee=report["employee"],
        department=report["department"],
        monthly=report["monthly"],
        summary=report["summary"],
        year=year
    )

@app.route("/leave-report/individual/<int:user_id>/download")
def download_individual_leave_pdf(user_id):
    year = request.args.get("year", datetime.now().year)

    report = build_individual_leave_report(user_id, year)

    html = render_template(
        "reports/individual_leave_report.html",
        report=report,
        year=year
    )

    pdf = HTML(string=html).write_pdf(
        stylesheets=[CSS(string="""
            body { font-family: Arial; font-size: 12px; }
        """)]
    )

    return Response(
        pdf,
        mimetype="application/pdf",
        headers={
            "Content-Disposition":
            f"attachment; filename=leave_report_{report['employee']['name']}.pdf"
        }
    )

@app.route("/leave-report/<int:user_id>/<string:format>")
def download_individual_leave_report(user_id, format):
    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        ("""SELECT
            u.full_name,
            d.name AS department_name,
            l.leave_type,
            l.start_date,
            l.end_date
        FROM leaves l
        JOIN users u ON u.id = l.user_id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE u.id = %s AND l.status = 'Approved'
        ORDER BY l.start_date
    """), 
        (user_id,)
    )
    
    rows = cur.fetchall()
    conn.close()

    if format == "excel":
        import pandas as pd
        df = pd.DataFrame(rows)
        output = io.BytesIO()
        df.to_excel(output, index=False)
        output.seek(0)

        return send_file(
            output,
            download_name="leave_report.xlsx",
            as_attachment=True
        )

    # ===== PDF =====
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)

    y = 800
    c.setFont("Helvetica-Bold", 14)
    c.drawString(50, y, f"Leave Report - {rows[0]['full_name']}")
    y -= 30

    c.setFont("Helvetica", 10)
    for r in rows:
        c.drawString(
            50, y,
            f"{r['start_date']} → {r['end_date']} | {r['leave_type']}"
        )
        y -= 18
        if y < 50:
            c.showPage()
            y = 800

    c.save()
    buffer.seek(0)

    return send_file(buffer, as_attachment=True, download_name="leave_report.pdf")

@app.route("/admin/dashboard/view/<filter>")
@admin_required
def admin_dashboard_view(filter):
    conn = get_db()
    c = conn.cursor()

    # ====== Ambil nilai filter dari query string ======
    filter_type     = request.args.get("type", "").strip()
    filter_position = request.args.get("position", "").strip()
    filter_alpha    = request.args.get("alpha", "").strip()     # A, B, C...
    date_from       = request.args.get("date_from", "").strip()
    date_to         = request.args.get("date_to", "").strip()

    # ====== Base query ikut "filter" dalam URL (today/month/pending...) ======
    base_query = "SELECT * FROM leave_applications WHERE 1=1 "
    params = []

    if filter == "month":
        title = "Leave Applications This Month"
        base_query += " AND EXTRACT(MONTH FROM start_date) = EXTRACT(MONTH FROM CURRENT_DATE) "

    elif filter == "today":
        title = "Employees On Leave Today"
        base_query += """
            AND status='Approved'
            AND CURRENT_DATE BETWEEN start_date AND end_date
        """

    elif filter == "pending":
        title = "Pending Leave Requests"
        base_query += " AND status IN ('Pending Recommender','Pending Approval') "

    # 🔹 TAMBAH BAHAGIAN INI
    elif filter == "approved":
        title = "Approved Leave Requests"
        base_query += " AND status='Approved' "

    elif filter == "rejected":
        title = "Rejected Leave Requests"
        base_query += " AND status='Rejected' "

    else:
        title = "Leave Details"

    # ====== Tambah FILTER TYPE (leave_type) ======
    if filter_type:
        base_query += " AND leave_type = %s "
        params.append(filter_type)

    # ====== Tambah FILTER POSITION ======
    if filter_position:
        base_query += " AND position = %s "
        params.append(filter_position)

    # ====== Tambah FILTER ALPHABET NAMA (full_name bermula dengan huruf) ======
    if filter_alpha:
        base_query += " AND full_name LIKE %s "
        params.append(f"{filter_alpha}%")

    # ====== Tambah FILTER TARIKH (range start_date) ======
    # date_from & date_to format: YYYY-MM-DD (HTML input type="date")
    if date_from and date_to:
        base_query += " AND start_date BETWEEN %s::date AND %s::date "
        params.extend([date_from, date_to])
    elif date_from:
        base_query += " AND start_date >= %s::date "
        params.append(date_from)
    elif date_to:
        base_query += " AND start_date <= %s::date "
        params.append(date_to)

    # Susun ikut tarikh
    base_query += " ORDER BY start_date DESC "

    c.execute((base_query), params)
    rows = c.fetchall()

    # ====== Data untuk dropdown filter (type & position) ======
    c.execute(
        ("SELECT DISTINCT leave_type FROM leave_applications WHERE leave_type IS NOT NULL"))

    rows = c.fetchall()
    leave_types = [r["leave_type"] for r in rows]

    c.execute(
        ("SELECT DISTINCT position FROM leave_applications WHERE position IS NOT NULL"))

    rows = c.fetchall()
    positions = [r["position"] for r in rows]

    conn.close()

    return render_template(
        "admin_dashboard_detail.html",
        title=title,
        leaves=rows,
        leave_types=leave_types,
        positions=positions,
        # untuk isi semula value dalam form
        filter_type=filter_type,
        filter_position=filter_position,
        filter_alpha=filter_alpha,
        date_from=date_from,
        date_to=date_to,
        current_filter=filter,
    )
    
@app.route("/admin/leave-report/view")
@admin_required
def view_all_leave_report():
    month = request.args.get("month")
    year = request.args.get("year")

    conn = get_db()
    c = conn.cursor()

    query = """
        SELECT
            la.full_name,
            d.name AS department,
            COUNT(*) AS total_applications,
            SUM(la.total_days) AS total_days
        FROM leave_applications la
        LEFT JOIN users u ON la.user_id=u.id
        LEFT JOIN departments d ON u.department_id=d.id
        WHERE la.status='Approved'
    """
    params = []

    if month:
        query += " AND TO_CHAR(la.start_date, 'YYYY-MM')=%s"
        params.append(month)
    elif year:
        query += " AND EXTRACT(YEAR FROM la.start_date)=%s"
        params.append(year)

    query += " GROUP BY la.user_id ORDER BY la.full_name"

    c.execute(query, params)
    rows = c.fetchall()
    conn.close()

    return render_template(
        "reports/all_leave_report.html",
        rows=rows,
        month=month,
        year=year
    )

def get_dashboard_data(date_from=None, date_to=None):
    conn = get_db()
    c = conn.cursor()

    # ===================== SUMMARY CARDS =======================
    c.execute(
        ("""SELECT
            COUNT(*) FILTER (
                WHERE EXTRACT(MONTH FROM start_date) =
                    EXTRACT(MONTH FROM CURRENT_DATE)
            ) AS total_this_month,

            COUNT(*) FILTER (
                WHERE status='Approved'
                AND CURRENT_DATE BETWEEN start_date AND end_date
            ) AS leave_today,

            COUNT(*) FILTER (
                WHERE status IN ('Pending Recommender','Pending Approval')
            ) AS pending_leave,

            COUNT(*) FILTER (
                WHERE status='Approved'
            ) AS approved_leave,

            COUNT(*) FILTER (
                WHERE status='Rejected'
            ) AS rejected_leave

        FROM leave_applications
    """))

    summary = c.fetchone()

    total_this_month = summary["total_this_month"]
    leave_today      = summary["leave_today"]
    pending_leave    = summary["pending_leave"]
    approved_leave   = summary["approved_leave"]
    rejected_leave   = summary["rejected_leave"]


    # ===================== RECENT REQUEST LIST ==================
    c.execute(
        ("""SELECT id, full_name, leave_type, status
        FROM leave_applications
        ORDER BY created_at DESC LIMIT 7
    """))
    recent_requests = c.fetchall()

    # ===================== ON LEAVE TODAY LIST ==================
    c.execute(
        ("""SELECT 
            la.full_name, 
            COALESCE(d.name, '-') AS department, 
            la.end_date AS return_date
        FROM leave_applications la
        LEFT JOIN users u ON u.id = la.user_id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE la.status = 'Approved'
        AND CURRENT_DATE BETWEEN la.start_date AND la.end_date
    """))
    on_leave_today = c.fetchall()

    # ===================== DEPARTMENTS FOR FILTER ================
    c.execute(
        ("SELECT DISTINCT name FROM departments WHERE name IS NOT NULL"))
    departments = [row["name"] for row in c.fetchall()]

    # ===================== PIE (LEAVE TYPES) =====================
    c.execute(
        ("""SELECT leave_type, COUNT(*) AS total
        FROM leave_applications
        GROUP BY leave_type
    """))
    rows = c.fetchall()
    leave_types = [r["leave_type"] for r in rows]
    leave_type_count = [r["total"] for r in rows]

    # ===================== TREND (WEEKLY) ========================
    c.execute(
        ("""SELECT 
            TO_CHAR(start_date, 'DD/MM') AS label,
            COUNT(*) AS total
        FROM leave_applications
        WHERE start_date >= CURRENT_DATE - INTERVAL '7 days'
        GROUP BY start_date
    """))

    t = c.fetchall()
    trend_labels = [r["label"] for r in t]
    trend_data   = [r["total"] for r in t]

    # ===================== MC RECORDS =====================
    c.execute(
        ("""SELECT 
            m.id,
            u.full_name,
            u.address,
            m.mc_number,
            m.start_date,
            m.end_date,
            m.pdf_path,
            m.created_at
        FROM mc_records m
        JOIN users u ON u.id = m.user_id
        ORDER BY m.created_at DESC
        LIMIT 10
    """))
    mc_records = c.fetchall()
    # ===================== MC COUNT PER MONTH =====================
    c.execute(
        ("""SELECT 
            TO_CHAR(created_at::date, 'MM/YYYY') AS label,
            COUNT(*) AS total
        FROM mc_records
        GROUP BY label
        ORDER BY label DESC
        LIMIT 6
    """))

    mc_rows = c.fetchall()
    mc_labels = [r["label"] for r in mc_rows]
    mc_counts = [r["total"] for r in mc_rows]

    # ===================== APPROVED / REJECTED DETAILS ===========
    range_clauses = []
    if date_from:
        range_clauses.append("start_date >= %s::date")
    if date_to:
        range_clauses.append("start_date <= %s::date")

    where_range = ""
    if range_clauses:
        where_range = " AND " + " AND ".join(range_clauses)

    # --- Rejected details ---
    params_rej = []
    if date_from:
        params_rej.append(date_from)
    if date_to:
        params_rej.append(date_to)

    c.execute(
        ("""SELECT id, full_name, leave_type, start_date, end_date, status
        FROM leave_applications
        WHERE status = 'Rejected' """ + where_range + """
        ORDER BY start_date DESC
        LIMIT 10
    """), params_rej
    )
    rejected_details = c.fetchall()

    # --- Approved details ---
    params_app = []
    if date_from:
        params_app.append(date_from)
    if date_to:
        params_app.append(date_to)

    c.execute(
        ("""SELECT id, full_name, leave_type, start_date, end_date, status
        FROM leave_applications
        WHERE status = 'Approved' """ + where_range + """
        ORDER BY start_date DESC
        LIMIT 10
    """), params_app
    )
    approved_details = c.fetchall()

    conn.close()

    return {
        "total_this_month": total_this_month,
        "leave_today": leave_today,
        "pending_leave": pending_leave,
        "approved_leave": approved_leave,
        "rejected_leave": rejected_leave,
        "recent_requests": recent_requests,
        "on_leave_today": on_leave_today,
        "departments": departments,
        "leave_types": leave_types,
        "leave_type_count": leave_type_count,
        "trend_labels": trend_labels,
        "trend_data": trend_data,
        "inc_month": 12,
        "diff_today": 3,
        "rejected_details": rejected_details,
        "approved_details": approved_details,
        "mc_records": mc_records,
        "mc_labels": mc_labels,
        "mc_counts": mc_counts,
    }

def get_leave_data(mode):
    conn = get_db()
    c = conn.cursor()

    if mode == "weekly":
        query = """
            SELECT u.full_name, l.leave_type, l.start_date, l.end_date, l.status
            FROM leaves l
            JOIN users u ON l.user_id = u.id
            WHERE l.start_date >= CURRENT_DATE - INTERVAL '7 days'
            ORDER BY l.start_date ASC
        """
    else:  # alphabetical
        query = """
            SELECT u.full_name, l.leave_type, l.start_date, l.end_date, l.status
            FROM leaves l
            JOIN users u ON l.user_id = u.id
            ORDER BY u.full_name ASC
        """

    c.execute(query)
    rows = c.fetchall()
    conn.close()
    return rows

@app.route("/admin/users")
@admin_required
def manage_users():
    conn = get_db()
    c = conn.cursor()

    today = date.today().isoformat()

    filter_name  = request.args.get("name", "")
    filter_dept  = request.args.get("department", "")
    filter_avail = request.args.get("availability", "")
    sort         = request.args.get("sort", "id_desc")

    base_query = """
        SELECT 
            u.id,
            u.username,
            u.full_name,
            u.role,
            u.position,
            u.entitlement,
            u.enrollment_date,
            u.created_at,
            u.department_id,
            u.ic_number,
            u.email,
            u.phone,
            u.address,

            CASE
                WHEN EXISTS (
                    SELECT 1 FROM leave_applications la
                    WHERE la.user_id = u.id
                    AND la.status = 'Approved'
                    AND %s::date BETWEEN la.start_date AND la.end_date
                ) THEN 'On Leave'
                ELSE COALESCE(u.availability, 'Available')
            END AS availability,

            d.name AS department_name

        FROM users u
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE 1=1
    """


    params = [today]

    if filter_name:
        base_query += " AND u.full_name LIKE %s"
        params.append(f"%{filter_name}%")

    if filter_dept:
        base_query += " AND d.name = %s"
        params.append(filter_dept)

    if filter_avail:
        base_query += """
            AND (
                CASE
                    WHEN EXISTS (
                        SELECT 1 FROM leave_applications la
                        WHERE la.user_id = u.id
                        AND la.status = 'Approved'
                        AND %s::date BETWEEN la.start_date AND la.end_date
                    ) THEN 'On Leave'
                    ELSE COALESCE(u.availability, 'Available')
                END
            ) = %s
        """
        params.extend([today, filter_avail])

    if sort == "az":
        base_query += " ORDER BY u.full_name ASC"
    elif sort == "za":
        base_query += " ORDER BY u.full_name DESC"
    else:
        base_query += " ORDER BY u.id DESC"

    c.execute((base_query), params)
    users = [dict(row) for row in c.fetchall()]

    c.execute(
        ("SELECT * FROM departments ORDER BY name"))
    departments = [dict(row) for row in c.fetchall()]

    conn.close()

    return render_template(
        "manage_users.html",
        users=users,
        departments=departments,
        current_user_role=session.get("role")
    )

@app.route("/admin/users/update_availability/<int:user_id>", methods=["POST"])
@admin_required
def update_availability(user_id):
    """
    AJAX endpoint:
    Expect JSON body: { "availability": "Available" | "Out" | "MC" | "WFH" }
    Returns JSON { success: True, prev: "<old>" } or error.
    """
    import json
    try:
        data = request.get_json(force=True)
    except Exception:
        return jsonify({"success": False, "error": "Invalid JSON"}), 400

    new_status = (data.get("availability") or "").strip()
    if new_status not in ("Available", "Out", "MC", "WFH", "Resign"):
        return jsonify({"success": False, "error": "Invalid availability"}), 400

    conn = get_db()
    c = conn.cursor()

    c.execute(
        ("SELECT availability FROM users WHERE id=%s"), 
        (user_id,)
    )
    row = c.fetchone()
    if not row:
        conn.close()
        return jsonify({"success": False, "error": "User not found"}), 404

    prev = row["availability"] if row["availability"] is not None else ""

    try:
        c.execute(
            ("UPDATE users SET availability=%s WHERE id=%s"), 
            (new_status, user_id)
        )
        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({"success": False, "error": f"DB error: {str(e)}"}), 500

    conn.close()
    return jsonify({"success": True, "prev": prev})

from flask import request, jsonify
from werkzeug.security import generate_password_hash

@app.route("/admin/users/reset_login/<int:user_id>", methods=["POST"])
@admin_required
def reset_login(user_id):

    data = request.get_json(force=True)

    new_username = data.get("username")
    new_password = data.get("password")

    if not new_username or not new_password:
        return jsonify({"success": False, "msg": "Missing data"}), 400

    hashed_pw = generate_password_hash(new_password)

    conn = get_db()
    c = conn.cursor()

    try:
        c.execute(
            ("""UPDATE users
                SET username=%s, password_hash=%s
                WHERE id=%s
            """), 
            (new_username, hashed_pw, user_id)
        )

        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({"success": False, "error": str(e)}), 500

    conn.close()
    return jsonify({"success": True})


@app.route("/admin/upload_mc", methods=["POST"])
@admin_required
def upload_mc():
    """
    Form fields expected:
      - user_id
      - mc_number (optional)
      - mc_start (optional, YYYY-MM-DD)
      - mc_end (optional)
      - mc_pdf (file, required)
    After saving file -> insert into mc_records, set users.availability='Out'
    """
    user_id = request.form.get("user_id")
    mc_number = request.form.get("mc_number", "").strip()
    mc_start = request.form.get("mc_start", "").strip() or None
    mc_end = request.form.get("mc_end", "").strip() or None
    pdf = request.files.get("mc_pdf")

    if not user_id:
        flash("Sila pilih pengguna untuk MC.", "warning")
        return redirect(url_for("manage_users"))

    # Validate user exists
    conn = get_db()
    c = conn.cursor()
    c.execute(
        ("SELECT id FROM users WHERE id=%s"), 
        (user_id,)
    )
    if not c.fetchone():
        conn.close()
        flash("Pengguna tidak ditemui.", "danger")
        return redirect(url_for("manage_users"))

    if not pdf or pdf.filename == "":
        conn.close()
        flash("Sila muat naik fail MC (PDF/JPG/PNG).", "warning")
        return redirect(url_for("manage_users"))

    # Validate extension
    filename = secure_filename(pdf.filename)
    ext = filename.rsplit(".", 1)[-1].lower()
    if ext not in ALLOWED_LEAVE_EXTENSIONS:
        conn.close()
        flash("Jenis fail tidak dibenarkan. Gunakan PDF/PNG/JPG.", "danger")
        return redirect(url_for("manage_users"))

    # Save file
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    save_name = f"mc_{user_id}_{timestamp}_{filename}"
    save_path = os.path.join(LEAVE_UPLOAD_FOLDER, save_name)
    try:
        pdf.save(save_path)
    except Exception as e:
        conn.close()
        flash("Gagal simpan fail: " + str(e), "danger")
        return redirect(url_for("manage_users"))

    # Insert MC record + set availability to Out
    try:
        c.execute(
            ("""INSERT INTO mc_records (user_id, mc_number, start_date, end_date, pdf_path, uploaded_by, created_at)
            VALUES (%s,%s,%s,%s,%s,%s,%s)
        """), (user_id, mc_number, mc_start, mc_end, save_name, session.get("user_id"), datetime.utcnow()))

        c.execute(
            ("UPDATE users SET availability=%s WHERE id=%s"), 
            ("Out", user_id)
        )
        conn.commit()
        flash("MC berjaya dimuat naik dan status pengguna dikemaskini.", "success")
    except Exception as e:
        conn.rollback()
        flash("Gagal simpan rekod MC: " + str(e), "danger")
    finally:
        conn.close()

    return redirect(url_for("manage_users"))

@app.route("/api/mc-trend")
@admin_required
def mc_trend_api():
    view = request.args.get("view", "monthly")  # weekly | monthly

    conn = get_db()
    c = conn.cursor()

    if view == "weekly":
        # 🗓️ Last 7 days (including today)
        c.execute(
            ("""SELECT 
                TO_CHAR(created_at::date, 'DD/MM') AS label,
                COUNT(*) AS total
            FROM mc_records
            WHERE created_at::date >= CURRENT_DATE - INTERVAL '6 days'
            GROUP BY created_at::date
            ORDER BY created_at::date
        """))
    else:
        # 🗓️ Last 6 months
        c.execute(
            ("""SELECT 
                TO_CHAR(created_at::date, 'MM/YYYY') AS label,
                COUNT(*) AS total
            FROM mc_records
            WHERE created_at::date >= CURRENT_DATE - INTERVAL '5 months'
            GROUP BY TO_CHAR(created_at::date, 'YYYY-MM')
            ORDER BY TO_CHAR(created_at::date, 'YYYY-MM')
        """))

    rows = c.fetchall()
    conn.close()

    labels = [r["label"] for r in rows]
    data   = [r["total"] for r in rows]

    return jsonify({
        "labels": labels,
        "data": data
    })


@app.route("/api/user/<int:user_id>")
@admin_required
def api_user(user_id):
    conn = get_db()
    c = conn.cursor()
    c.execute(
        ("""SELECT 
            u.*,
            d.name AS department_name
        FROM users u
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE u.id = %s
    """), 
        (user_id,)
    )
    user = c.fetchone()
    conn.close()

    if not user:
        return jsonify({"error": "User not found"}), 404

    return jsonify(dict(user))

@app.route("/admin/users/create", methods=["POST"])
@admin_required
def create_user():
    username        = request.form.get("username", "").strip()
    full_name       = request.form.get("full_name", "").strip()
    role            = request.form.get("role", "user")
    password        = request.form.get("password", "")
    department_id   = request.form.get("dept_id") or None
    position        = request.form.get("position", "").strip()
    enrollment_date = request.form.get("enrollment_date", "").strip()
    entitlement     = request.form.get("entitlement", "").strip()

    # extra fields
    email      = request.form.get("email", "").strip()
    phone      = request.form.get("phone", "").strip()
    ic_number  = request.form.get("ic_number", "").strip()
    address    = request.form.get("address", "").strip()

    if not username or not full_name or not password:
        flash("All fields are required.", "danger")
        return redirect(url_for("manage_users"))

    try:
        entitlement_val = int(entitlement) if entitlement != "" else 0
    except ValueError:
        entitlement_val = 0

    conn = get_db()
    c = conn.cursor()
    try:
        c.execute(
            ("""INSERT INTO users (
                username,
                full_name,
                password_hash,
                role,
                created_at,
                department_id,
                position,
                enrollment_date,
                entitlement,
                email,
                phone,
                ic_number,
                address
            )
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """), (
            username,
            full_name,
            generate_password_hash(password),
            role,
            datetime.utcnow(),
            department_id,
            position,
            enrollment_date,
            entitlement_val,
            email,
            phone,
            ic_number,
            address
        )
    )
        conn.commit()
        flash("User created successfully.", "success")
    except Exception:
        flash("Username already exists.", "danger")
    finally:
        conn.close()
    return redirect(url_for("manage_users"))

@app.route("/admin/users/delete/<int:user_id>", methods=["POST"])
@admin_required
def delete_user(user_id):
    if user_id == session.get("user_id"):
        flash("You cannot delete yourself.", "warning")
        return redirect(url_for("manage_users"))

    conn = get_db()
    c = conn.cursor()
    c.execute(
        ("DELETE FROM users WHERE id=%s"), 
        (user_id,)
    )
    conn.commit()
    conn.close()

    flash("User deleted.", "success")
    return redirect(url_for("manage_users"))


@app.route("/admin/users/update/<int:user_id>", methods=["POST"])
@admin_required
def update_user_details(user_id):

    conn = get_db()
    c = conn.cursor()

    entitlement = request.form.get("entitlement")
    entitlement = int(entitlement) if entitlement and entitlement.isdigit() else 0

    availability_value = request.form.get("availability")

    dept_id = request.form.get("dept_id")

    # ✅ if admin did NOT select new department, keep old one
    if not dept_id:
        c.execute(
            ("SELECT department_id FROM users WHERE id=%s"), 
            (user_id,)
        )
        row = c.fetchone()
        dept_id = row["department_id"] if row else None

    c.execute(
        ("""UPDATE users SET 
            full_name     = %s,
            email         = %s,
            phone         = %s,
            address       = %s,
            position      = %s,
            department_id = %s,
            availability  = %s,
            entitlement   = %s
        WHERE id = %s
    """), (
        request.form.get("full_name"),
        request.form.get("email"),
        request.form.get("phone"),
        request.form.get("address"),
        request.form.get("position"),
        dept_id,                
        availability_value,
        entitlement,
        user_id
    ))

    conn.commit()
    conn.close()

    flash("User updated successfully.", "success")
    return "", 204


@app.route("/admin/users/entitlement/<int:user_id>", methods=["POST"])
@admin_required
def update_entitlement(user_id):
    """<<< This fixes the missing endpoint used by the template >>>"""
    raw = request.form.get("entitlement", "0").strip()
    try:
        value = max(0, int(raw))
    except ValueError:
        flash("Entitlement must be a whole number.", "danger")
        return redirect(url_for("manage_users"))

    conn = get_db()
    c = conn.cursor()
    c.execute(
        ("UPDATE users SET entitlement=%s WHERE id=%s"), 
        (value, user_id)
    )
    conn.commit()
    conn.close()
    flash("Entitlement updated.", "success")
    return redirect(url_for("manage_users"))

@app.route("/manage_leaves")
def manage_leaves():
    from datetime import datetime, timedelta

    conn = get_db()
    cur = conn.cursor()

    current_year = datetime.now().year
    start_year = current_year - 5
    year_range = list(range(start_year, current_year + 1))

    # =======================
    # LEAVE REPORT FILTERS
    # =======================
    report_year = request.args.get("year", str(current_year))
    selected_department = request.args.get("department", "all")
    action = request.args.get("action")

    should_build_report = action == "filter"

    # =======================
    # MONTHLY MATRIX FILTERS
    # =======================
    selected_month = request.args.get("matrix_month", datetime.now().strftime("%m"))
    matrix_year = request.args.get("matrix_year", str(current_year))
    selected_dept = request.args.get("matrix_department", "all")

    # =======================
    # DEPARTMENTS
    # =======================
    cur.execute(
        ("SELECT id, name FROM departments ORDER BY name"))
    departments = cur.fetchall()

    # =======================
    # LEAVE REPORT
    # =======================
    leave_report = []

    MONTH_MAP = {
        "01":"JAN","02":"FEB","03":"MAR","04":"APR",
        "05":"MAY","06":"JUN","07":"JUL","08":"AUG",
        "09":"SEP","10":"OCT","11":"NOV","12":"DEC"
    }

    if should_build_report:
        params = [report_year]
        dept_filter = ""

        if selected_department != "all":
            dept_filter = "AND d.name = %s"
            params.append(selected_department)

        cur.execute(
            (f"""SELECT u.id AS user_id,
                u.full_name,
                la.leave_type,
                la.start_date,
                la.end_date,
                u.entitlement
            FROM leave_applications la
            JOIN users u ON u.id = la.user_id
            LEFT JOIN departments d ON u.department_id = d.id
            WHERE la.status = 'Approved'
            AND EXTRACT(YEAR FROM la.start_date) = %s
            {dept_filter}
            ORDER BY u.full_name, la.start_date
        """), params
        )

        rows = cur.fetchall()
        users = {}

        for r in rows:
            uid = r["user_id"]

            # ===== INIT USER (FIXED ENTITLEMENT) =====
            if uid not in users:
                users[uid] = {
                    "user_id": uid,
                    "name": r["full_name"],
                    "entitlement": r["entitlement"] or 0,
                    "total_used": 0,
                    "monthly": {m: 0 for m in MONTH_MAP.values()},
                    "monthly_details": {},
                    "leave_type_details": {}
                }

            if r["leave_type"] == "MC":
                continue

            # ===== CALCULATE WORKING DAYS =====
            days = calculate_working_days(r["start_date"], r["end_date"])

            start = datetime.strptime(r["start_date"], "%Y-%m-%d")
            m = MONTH_MAP[start.strftime("%m")]

            # ===== ACCUMULATE USED =====
            users[uid]["monthly"][m] += days
            users[uid]["total_used"] += days

            # ===== MONTHLY DETAILS =====
            users[uid]["monthly_details"].setdefault(m, {})
            users[uid]["monthly_details"][m].setdefault(r["leave_type"], [])
            users[uid]["monthly_details"][m][r["leave_type"]].append({
                "start": r["start_date"],
                "end": r["end_date"],
                "days": days
            })

            # ===== LEAVE TYPE DETAILS =====
            users[uid]["leave_type_details"].setdefault(r["leave_type"], [])
            users[uid]["leave_type_details"][r["leave_type"]].append({
                "start": r["start_date"],
                "end": r["end_date"],
                "days": days
            })

        # ===== FINAL BALANCE CALCULATION (SAFE) =====
        for u in users.values():
            u["remaining"] = max(0, u["entitlement"] - u["total_used"])

        leave_report = list(users.values())

    # =========================================================
    # MONTHLY LEAVE MATRIX (FIXED WITH MC)
    # =========================================================
    monthly_matrix = []

    first_day = datetime.strptime(
        f"{matrix_year}-{selected_month}-01", "%Y-%m-%d"
    )

    if selected_month == "12":
        last_day = datetime.strptime(
            f"{int(matrix_year)+1}-01-01", "%Y-%m-%d"
        ) - timedelta(days=1)
    else:
        last_day = datetime.strptime(
            f"{matrix_year}-{int(selected_month)+1:02d}-01", "%Y-%m-%d"
        ) - timedelta(days=1)

    users = {}

    params = [last_day.strftime("%Y-%m-%d"), first_day.strftime("%Y-%m-%d")]
    dept_filter = ""

    if selected_dept != "all":
        dept_filter = "AND d.name = %s"
        params.append(selected_dept)

    # ===== APPROVED LEAVES =====
    cur.execute(
        (f"""SELECT
            la.user_id,
            u.full_name,
            la.leave_type,
            la.start_date,
            la.end_date
        FROM leave_applications la
        JOIN users u ON u.id = la.user_id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE la.status='Approved'
        AND la.start_date <= %s
        AND la.end_date >= %s
        {dept_filter}
        ORDER BY u.full_name
    """), params
    )

    rows = cur.fetchall()

    for r in rows:
        uid = r["user_id"]
        users.setdefault(uid, {
            "user_name": r["full_name"],
            "leaves": {}
        })

        start = datetime.strptime(r["start_date"], "%Y-%m-%d")
        end   = datetime.strptime(r["end_date"], "%Y-%m-%d")

        cur_day = max(start, first_day)
        while cur_day <= min(end, last_day):
            users[uid]["leaves"][cur_day.strftime("%d")] = r["leave_type"]
            cur_day += timedelta(days=1)


    # ===== MC RECORDS (FIXED) =====
    params = [last_day.strftime("%Y-%m-%d"), first_day.strftime("%Y-%m-%d")]
    dept_filter = ""

    if selected_dept != "all":
        dept_filter = "AND d.name = %s"
        params.append(selected_dept)

    cur.execute(
        (f"""SELECT
            m.user_id,
            u.full_name,
            m.start_date,
            m.end_date
        FROM mc_records m
        JOIN users u ON u.id = m.user_id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE m.start_date <= %s
        AND m.end_date>= %s
        {dept_filter}
    """), params
    )

    mc_rows = cur.fetchall()

    for m in mc_rows:
        uid = m["user_id"]
        users.setdefault(uid, {
            "user_name": m["full_name"],
            "leaves": {}
        })

        start = datetime.strptime(m["start_date"], "%Y-%m-%d")
        end   = datetime.strptime(m["end_date"], "%Y-%m-%d")

        cur_day = max(start, first_day)
        while cur_day <= min(end, last_day):
            users[uid]["leaves"][cur_day.strftime("%d")] = "MC"
            cur_day += timedelta(days=1)

    monthly_matrix = list(users.values())


    # =======================
    # COMPLETED LIST
    # =======================
    cur.execute(
        ("""
        SELECT la.id, la.full_name,
               COALESCE(d.name,'') AS department_name,
               la.leave_type, la.start_date, la.end_date,
               la.status, la.approver_name
        FROM leave_applications la
        LEFT JOIN users u ON u.id=la.user_id
        LEFT JOIN departments d ON u.department_id=d.id
        WHERE la.status IN ('Approved','Rejected')
        ORDER BY la.start_date DESC
    """))
    completed = cur.fetchall()

    conn.close()

    return render_template(
        "manage_leaves.html",
        leave_report=leave_report,
        selected_year=report_year,
        selected_department=selected_department,

        monthly_matrix=monthly_matrix,
        selected_month=selected_month,
        selected_year_matrix=matrix_year,
        selected_dept=selected_dept,

        completed=completed,
        departments=departments,
        year_range=year_range
    )


MONTHS = ["JAN","FEB","MAR","APR","MAY","JUN","JUL","AUG","SEP","OCT","NOV","DEC"]

from weasyprint import HTML
from flask import make_response
from datetime import date

@app.route("/completed-leaves/pdf")
@admin_required
def completed_leaves_pdf():

    name = request.args.get("name","")
    dept = request.args.get("dept","")
    year = request.args.get("year","")
    month = request.args.get("month","")
    status = request.args.get("status","")

    conn = get_db()
    cur = conn.cursor()

    query = """
        SELECT la.id, la.full_name,
               COALESCE(d.name,'') AS department_name,
               la.leave_type, la.start_date, la.end_date,
               la.status, la.approver_name
        FROM leave_applications la
        LEFT JOIN users u ON u.id=la.user_id
        LEFT JOIN departments d ON u.department_id=d.id
        WHERE la.status IN ('Approved','Rejected')
    """
    params = []

    if name:
        query += " AND la.full_name LIKE %s"
        params.append(name + "%")

    if dept:
        query += " AND d.name = %s"
        params.append(dept)

    if year:
        query += " AND EXTRACT(YEAR FROM la.start_date)  = %s"
        params.append(year)

    if month:
        query += " AND EXTRACT(MONTH FROM la.start_date) = %s"
        params.append(month)

    if status:
        query += " AND la.status = %s"
        params.append(status)

    query += " ORDER BY la.start_date DESC"

    cur.execute((query), params)
    completed = cur.fetchall()
    conn.close()

    html = render_template(
        "reports/completed_leaves_pdf.html",
        completed=completed,
        printed_date=date.today().strftime("%d-%m-%Y")
    )

    pdf = HTML(string=html).write_pdf()

    response = make_response(pdf)
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = "attachment; filename=Completed_Leaves_List.pdf"
    return response

def get_leave_matrix_report(year, department_id=None, user_id=None):
    conn = get_db()
    c = conn.cursor()

    sql = """
        SELECT
            u.id AS user_id,
            u.full_name,
            la.leave_type,
            CAST(EXTRACT(MONTH FROM la.start_date) AS INTEGER) AS month_no,
            SUM(la.total_days) AS days_used,
            u.entitlement
        FROM leave_applications la
        JOIN users u ON u.id = la.user_id
        WHERE la.status = 'Approved'
        AND EXTRACT(YEAR FROM la.start_date)  = %s
    """
    params = [year]

    if department_id:
        sql += " AND u.department_id = %s"
        params.append(department_id)

    if user_id:
        sql += " AND u.id = %s"
        params.append(user_id)

    sql += """
        GROUP BY u.id, la.leave_type, month_no
        ORDER BY u.full_name, la.leave_type
    """

    c.execute((sql), params)
    rows = c.fetchall()

    conn.close()

    return rows

def build_leave_matrix(rows):
    report = {}

    for r in rows:
        uid = r["user_id"]

        if uid not in report:
            report[uid] = {
                "user_id": uid,
                "name": r["full_name"],
                "entitlement": r["entitlement"] or 0,
                "leaves": []
            }

        leave = next(
            (l for l in report[uid]["leaves"] if l["leave_type"] == r["leave_type"]),
            None
        )

        if not leave:
            leave = {
                "leave_type": r["leave_type"],
                "months": {m: 0.0 for m in MONTHS},
                "total_used": 0.0
            }
            report[uid]["leaves"].append(leave)

        month_name = MONTHS[r["month_no"] - 1]
        leave["months"][month_name] += r["days_used"]
        leave["total_used"] += r["days_used"]

    # ✅ FINAL CALCULATION (SAFE)
    for u in report.values():
        used = sum(l["total_used"] for l in u["leaves"])
        u["remaining"] = max(0, u["entitlement"] - used)

    return list(report.values())


@app.route("/download/leave-report/excel")
def download_leave_report_excel():
    import pandas as pd
    from io import BytesIO
    from flask import send_file

    year = request.args.get("year")
    department = request.args.get("department", "all")

    conn = get_db()
    c = conn.cursor()

    query = """
      SELECT 
        u.full_name,
        d.name AS department,
        COALESCE(
          SUM(
            (l.end_date - l.start_date) + 1
        ), 0
        ) AS used
      FROM users u
      LEFT JOIN departments d ON u.department_id = d.id
      LEFT JOIN leaves l
        ON l.user_id = u.id
       AND l.status = 'Approved'
       AND EXTRACT(YEAR FROM l.start_date)  = %s
    """

    params = [year]

    if department != "all":
        query += " WHERE d.name = %s"
        params.append(department)

    query += """
      GROUP BY u.id
      ORDER BY u.full_name
    """

    c.execute((query), params)
    rows = c.fetchall()

    df = pd.DataFrame(rows, columns=[
        "Employee Name",
        "Department",
        "Total Leave Used (Days)"
    ])

    output = BytesIO()
    df.to_excel(output, index=False, sheet_name="Leave Report")
    output.seek(0)

    filename = f"LEAVE_REPORT_{department.upper()}_{year}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename
    )

@app.route("/download/leave-report/pdf")
def download_leave_report_pdf():
    from datetime import date

    year = request.args.get("year")
    department = request.args.get("department", "all")

    conn = get_db()
    c = conn.cursor()

    query = """
      SELECT 
        u.full_name,
        d.name AS department,
        COALESCE(
            SUM(
            (l.end_date - l.start_date) + 1
            )
            , 0
        ) AS used
      FROM users u
      LEFT JOIN departments d ON u.department_id = d.id
      LEFT JOIN leaves l
        ON l.user_id = u.id
       AND l.status = 'Approved'
       AND EXTRACT(YEAR FROM l.start_date)  = %s
    """
    params = [year]

    if department != "all":
        query += " WHERE d.name = %s"
        params.append(department)

    query += """
      GROUP BY u.id
      ORDER BY u.full_name
    """

    c.execute(query, params)
    rows = c.fetchall()


    return render_template(
        "leave_report_department_pdf.html",
        year=year,
        department=department,
        printed_date=date.today().strftime("%d-%m-%Y"),
        rows=rows
    )


@app.route("/admin/leave-report/matrix")
@admin_required
def view_leave_matrix():
    year = request.args.get("year", datetime.now().year)
    department = request.args.get("department_id")

    rows = get_leave_matrix_report(year, department)
    matrix = build_leave_matrix(rows)

    return render_template(
        "reports/leave_matrix_preview.html",
        year=year,
        matrix=matrix,
        months=MONTHS
    )

@app.route("/admin/leave-report/matrix/download/excel")
@admin_required
def download_leave_matrix_excel():
    year = request.args.get("year", datetime.now().year)

    rows = get_leave_matrix_report(year)
    data = build_leave_matrix(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Leave Report"

    headers = ["ID","Name","Leave Type"] + MONTHS + ["Total Used","Remaining"]
    ws.append(headers)

    for r in data:
        ws.append([
            r["user_id"],
            r["name"],
            r["leave_type"],
            *[r["months"][m] for m in MONTHS],
            r["total_used"],
            r["remaining"]
        ])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return send_file(
        stream,
        as_attachment=True,
        download_name=f"Leave_Report_{year}.xlsx"
    )

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors

@app.route("/admin/leave-report/matrix/download/pdf")
@admin_required
def download_leave_matrix_pdf():
    year = request.args.get("year", datetime.now().year)

    rows = get_leave_matrix_report(year)
    data = build_leave_matrix(rows)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer)

    table_data = [["ID","Name","Leave Type"] + MONTHS + ["Total","Remaining"]]

    for r in data:
        table_data.append(
            [r["user_id"], r["name"], r["leave_type"]]
            + [r["months"][m] for m in MONTHS]
            + [r["total_used"], r["remaining"]]
        )

    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("ALIGN", (3,1), (-1,-1), "CENTER"),
    ]))

    doc.build([table])
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"Leave_Report_{year}.pdf",
        mimetype="application/pdf"
    )

def get_leave_report(report_type="monthly", month=None):
    conn = get_db()
    c = conn.cursor()

    if month:
        where_clause = "TO_CHAR(la.start_date, 'YYYY-MM')  = %s"
        params = [month]
    elif report_type == "yearly":
        where_clause = "EXTRACT(YEAR FROM la.start_date) = EXTRACT(YEAR FROM CURRENT_DATE)"
        params = []
    else:
        where_clause = "TO_CHAR(la.start_date, 'YYYY-MM') = TO_CHAR(CURRENT_DATE, 'YYYY-MM')"
        params = []

    c.execute(
        (f"""SELECT
            la.user_id,
            la.full_name,
            COALESCE(d.name, '-') AS department_name,
            COUNT(la.id) AS total_application,
            COALESCE(SUM(la.total_days), 0) AS total_days
        FROM leave_applications la
        LEFT JOIN users u ON u.id = la.user_id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE {where_clause}
        GROUP BY la.user_id
        ORDER BY la.full_name
    """), params
    )

    rows = c.fetchall()
    conn.close()
    return rows

def get_all_employee_leaves(view="monthly"):
    conn = get_db()
    c = conn.cursor()

    if view == "weekly":
        where_clause = "start_date >= CURRENT_DATE - INTERVAL '7 days'"
    else:  # monthly
        where_clause = "TO_CHAR(start_date, 'YYYY-MM') = TO_CHAR(CURRENT_DATE, 'YYYY-MM')"

    c.execute(
        (f"""SELECT
            la.id,
            la.full_name,
            COALESCE(d.name, '-') AS department_name,
            la.leave_type,
            la.start_date,
            la.end_date,
            la.status
        FROM leave_applications la
        LEFT JOIN users u ON u.id = la.user_id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE {where_clause}
        ORDER BY la.start_date DESC
    """))

    rows = c.fetchall()
    conn.close()
    return rows

@app.route("/admin/leaves/approve/<int:leave_id>", methods=["POST"])
@admin_required
def assign_approver(leave_id):
    """Admin chooses who will approve a leave after it’s checked."""
    approver_position = request.form.get("approver_position")
    approver_department = request.form.get("approver_department")

    if not approver_position:
        flash("Please select approver position.", "warning")
        return redirect(url_for("manage_leaves"))

    conn = get_db()
    c = conn.cursor()
    c.execute(
        ("""UPDATE leaves
        SET next_approver=%s, next_approver_position=%s, next_approver_department=%s
        WHERE id=%s
    """), 
        (approver_position, approver_position, approver_department, leave_id)
    )
    conn.commit()
    conn.close()

    flash(f"Leave assigned to be approved by {approver_position} ({approver_department or 'All Departments'}).", "info")
    return redirect(url_for("manage_leaves"))


@app.route("/leave/action/<int:leave_id>/<action>", methods=["POST"])
@login_required
def update_leave_status(leave_id, action):
    user_id = session["user_id"]

    conn = get_db()
    c = conn.cursor()

    # Fetch leave
    c.execute(
        ("SELECT * FROM leaves WHERE id=%s"), 
        (leave_id,)
    )
    leave = c.fetchone()

    if not leave:
        conn.close()
        flash("Leave not found.", "danger")
        return redirect(url_for("user_dashboard"))

    # --- ONLY ASSIGNED APPROVER CAN APPROVE ---
    if action == "approve":
        if leave["approved_by_user_id"] != user_id:
            flash("You are NOT authorized to approve this leave.", "danger")
            conn.close()
            return redirect(url_for("user_dashboard"))

        # Approve leave
        c.execute(
            ("UPDATE leaves SET status='Approved' WHERE id=%s"), 
            (leave_id,)
        )
        conn.commit()

        flash("Leave approved successfully.", "success")
        conn.close()
        return redirect(url_for("user_dashboard"))

    # --- ONLY CHECKER CAN CHECK OR REJECT ---
    if action == "check":
        if leave["checked_by_user_id"] != user_id:
            flash("You are NOT authorized to check this leave.", "danger")
            conn.close()
            return redirect(url_for("user_dashboard"))

        c.execute(
            ("UPDATE leaves SET checked_status='Checked' WHERE id=%s"), 
            (leave_id,)
        )
        conn.commit()

        flash("Leave checked successfully.", "success")
        conn.close()
        return redirect(url_for("user_dashboard"))

    if action == "reject":
        # Checker OR Approver can reject (depending on position)
        if leave["checked_by_user_id"] != user_id and leave["approved_by_user_id"] != user_id:
            flash("You are NOT authorized to reject this leave.", "danger")
            conn.close()
            return redirect(url_for("user_dashboard"))

        c.execute(
            ("UPDATE leaves SET status='Rejected' WHERE id=%s"), 
            (leave_id,)
        )
        conn.commit()

        flash("Leave rejected.", "info")
        conn.close()
        return redirect(url_for("user_dashboard"))


@app.route('/update-leave-status', methods=['POST'])
@login_required
def update_leave_status_modal():
    """AJAX endpoint for approving/rejecting leave directly from modal."""
    data = request.get_json()
    leave_id = data.get('id')
    new_status = data.get('status')

    if not leave_id or new_status not in ['Approved', 'Rejected', 'Pending']:
        return jsonify({'success': False, 'error': 'Invalid request'}), 400

    conn = get_db()
    c = conn.cursor()
    c.execute(
        ("SELECT id FROM leaves WHERE id=%s"), 
        (leave_id,)
    )
    leave = c.fetchone()

    if not leave:
        conn.close()
        return jsonify({'success': False, 'error': 'Leave not found'}), 404

    c.execute(
        ("UPDATE leaves SET status=%s WHERE id=%s"), 
        (new_status, leave_id)
    )
    conn.commit()
    conn.close()

    return jsonify({'success': True, 'status': new_status})


@app.route("/admin/holidays", methods=["GET", "POST"])
@admin_required
def holidays():
    conn = get_db()
    c = conn.cursor()
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        day = request.form.get("date", "").strip()
        if name and day:
            c.execute(
                ("INSERT INTO holidays (name, date) VALUES (%s,%s)"), 
                (name, day)
            )
            conn.commit()
            flash("Holiday added.", "success")
        else:
            flash("Please provide name and date.", "danger")
    c.execute(
        ("SELECT * FROM holidays ORDER BY date"))
    rows = c.fetchall()
    conn.close()
    return render_template("holidays.html", holidays=rows)

@app.route("/admin/departments", methods=["GET", "POST"])
@admin_required
def manage_departments():
    conn = get_db()
    c = conn.cursor()

    if request.method == "POST":
        dept_name = request.form.get("name", "").strip()
        if dept_name:
            try:
                c.execute(
                    ("INSERT INTO departments (name) VALUES (%s)"), 
                    (dept_name,)
                )
                conn.commit()
                flash("Department added successfully.", "success")
            except Exception:
                flash("Department already exists.", "warning")
            finally:
                conn.close()

            return redirect(url_for("manage_users"))

    # GET method (only if user visits /admin/departments directly)
    c.execute(
        ("SELECT * FROM departments ORDER BY name"))
    departments = c.fetchall()
    conn.close()

    return render_template("manage_departments.html", departments=departments)


@app.route("/admin/departments/delete/<int:dept_id>", methods=["POST"])
@admin_required
def delete_department(dept_id):
    """Delete department and stay on Manage Users dashboard."""
    conn = get_db()
    c = conn.cursor()
    c.execute(
        ("DELETE FROM departments WHERE id=%s"), 
        (dept_id,)
    )
    conn.commit()
    conn.close()
    flash("Department deleted successfully.", "info")

    # ✅ Stay on Manage Users dashboard
    return redirect(url_for("manage_users"))

@app.route("/ceo/dashboard")
@login_required
def ceo_dashboard():

    pos = (session.get("position") or "").strip().upper()
    if pos != "CEO":
        flash("Not authorized.", "danger")
        return redirect(url_for("user_dashboard"))

    conn = get_db()
    cur = conn.cursor()

    # ================== PENDING (CEO only) ==================
    cur.execute(
        ("""SELECT l.*, u.full_name, u.position AS user_position,
               d.name AS department_name
        FROM leave_applications l
        JOIN users u ON u.id = l.user_id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE l.status = 'Pending Approval'
          AND UPPER(TRIM(l.approver_name)) = 'CEO'
        ORDER BY l.id DESC
    """))
    pending_leaves = cur.fetchall()
    pending_count = len(pending_leaves)

    # ================== APPROVED (CEO only) ==================
    cur.execute(
        ("""SELECT l.*, u.full_name, u.position AS user_position,
               d.name AS department_name
        FROM leave_applications l
        JOIN users u ON u.id = l.user_id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE l.status = 'Approved'
          AND UPPER(TRIM(l.approver_name)) = 'CEO'
        ORDER BY l.approved_at DESC
    """))
    approved_leaves = cur.fetchall()
    approved_count = len(approved_leaves)

    # ================== REJECTED (CEO only) ==================
    cur.execute(
        ("""SELECT l.*, u.full_name, u.position AS user_position,
               d.name AS department_name
        FROM leave_applications l
        JOIN users u ON u.id = l.user_id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE l.status = 'Rejected'
          AND UPPER(TRIM(l.approver_name)) = 'CEO'
        ORDER BY l.approved_at DESC
    """))
    rejected_leaves = cur.fetchall()
    rejected_count = len(rejected_leaves)

    # ================== BAR GRAPH (Approved + Rejected per Month) ==================
    cur.execute(
        ("""SELECT 
            EXTRACT(YEAR FROM l.approved_at::date) AS year,
            EXTRACT(MONTH FROM l.approved_at::date) AS month,
            COALESCE(d.name,'Unknown') AS department,
            l.leave_type,
            SUM(CASE WHEN l.status='Approved' THEN 1 ELSE 0 END) AS approved,
            SUM(CASE WHEN l.status='Rejected' THEN 1 ELSE 0 END) AS rejected
        FROM leave_applications l
        JOIN users u ON u.id = l.user_id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE l.status IN ('Approved','Rejected')
        GROUP BY year, month, department, l.leave_type
        ORDER BY year, month
    """))
    stats = [dict(row) for row in cur.fetchall()]

    years = sorted({row["year"] for row in stats})

    # ================== TREND RAW DATA (for filters) ==================
    cur.execute(
        ("""SELECT 
            EXTRACT(YEAR FROM l.approved_at::date) AS year,
            EXTRACT(MONTH FROM l.approved_at::date) AS month,
            COALESCE(d.name,'Unknown') AS department,
            COUNT(*) AS total
        FROM leave_applications l
        JOIN users u ON u.id = l.user_id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE l.status IN ('Approved','Rejected')
        GROUP BY year, month, department
        ORDER BY year, month
    """))
    trend_raw = [dict(row) for row in cur.fetchall()]

    # ================== PIE CHART (Leave Type Distribution) ==================
    cur.execute(
        ("""SELECT 
            l.leave_type,
            COUNT(*) AS total
        FROM leave_applications l
        WHERE l.status IN ('Approved','Rejected')
        GROUP BY l.leave_type
    """))
    pie_data = cur.fetchall()

    leave_types = [row["leave_type"] for row in pie_data]
    leave_type_count = [row["total"] for row in pie_data]

    conn.close()

    return render_template(
        "ceo_dashboard.html",
        pending_leaves=pending_leaves,
        approved_leaves=approved_leaves,
        rejected_leaves=rejected_leaves,
        pending_count=pending_count,
        approved_count=approved_count,
        rejected_count=rejected_count,
        stats=stats,
        years=years,
        trend_raw=trend_raw,
        leave_types=leave_types,
        leave_type_count=leave_type_count
    )

@app.route("/ceo/leave/approve/<int:leave_id>", methods=["POST"])
@login_required
def ceo_approve_leave(leave_id):

    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        ("SELECT * FROM leave_applications WHERE id=%s"), 
        (leave_id,)
    )
    leave = cur.fetchone()

    cur.execute(
        ("""UPDATE leave_applications
        SET status='Approved',
            approved_at=%s
        WHERE id=%s
    """), (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), leave_id))

    if leave["leave_type"] != "MC":
        cur.execute(
            ("""UPDATE users
            SET entitlement = entitlement - %s
            WHERE id = %s
        """), 
            (leave["total_days"], leave["user_id"])
        )

    conn.commit()
    conn.close()

    flash("Leave approved by CEO. Balance updated.", "success")
    return redirect(url_for("ceo_dashboard"))

@app.route("/ceo/leave/reject/<int:leave_id>", methods=["POST"])
@login_required
def ceo_reject_leave(leave_id):

    if (session.get("position") or "").upper() != "CEO":
        flash("Not authorized.", "danger")
        return redirect(url_for("user_dashboard"))

    conn = get_db()
    cur = conn.cursor()

    cur.execute(("""UPDATE leave_applications
        SET status='Rejected',
            approved_at=%s
        WHERE id=%s
    """), (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), leave_id))

    conn.commit()
    conn.close()

    flash("Leave rejected by CEO.", "info")
    return redirect(url_for("ceo_dashboard"))


# ---------------------- User Views ----------------------
@app.route("/user/dashboard")
@login_required
def user_dashboard():

    user_id = session["user_id"]

    conn = get_db()
    c = conn.cursor()

    # 🟦 Get and normalize user position
    c.execute(
        ("SELECT position FROM users WHERE id=%s"), 
        (user_id,)
    )
    row = c.fetchone()
    my_pos = normalize(row["position"]) if row else ""
    
    # ===== ENTITLEMENT SUMMARY =====
    c.execute(
        ("SELECT entitlement FROM users WHERE id=%s"), 
        (user_id,)
    )
    entitlement = c.fetchone()["entitlement"] or 0

    c.execute(
        ("""SELECT start_date, end_date
        FROM leave_applications
        WHERE user_id = %s
        AND status = 'Approved'
        AND leave_type != 'MC'
    """), 
        (user_id,)
    )

    rows = c.fetchall()

    used = 0
    for r in rows:
        used += calculate_working_days(r["start_date"], r["end_date"])

    balance = max(0, entitlement - used)


    # ===================== MY OWN LEAVES =====================
    c.execute(
        ("""SELECT l.*, u.full_name, d.name AS department_name
        FROM leave_applications l
        JOIN users u ON u.id = l.user_id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE l.user_id=%s
        ORDER BY l.id DESC
    """), 
        (user_id,)
    )
    my_leaves = c.fetchall()

    # ===================== MY MC RECORDS =====================
    c.execute(
        ("""SELECT mc_number, start_date, end_date, pdf_path, created_at
        FROM mc_records
        WHERE user_id = %s
        ORDER BY created_at DESC
    """), 
        (user_id,)
    )
    my_mc = c.fetchall()

    # ===================== PENDING ACTIONS (NEW LOGIC) =====================
    c.execute(
        ("""SELECT 
            l.*, u.full_name, u.position, d.name AS department_name
        FROM leave_applications l
        JOIN users u ON u.id = l.user_id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE 
            (
                l.status = 'Pending Recommender'
                AND UPPER(TRIM(l.checker_name)) = %s
            )
            OR
            (
                l.status = 'Pending Approval'
                AND UPPER(TRIM(l.approver_name)) = %s
            )
        ORDER BY l.id DESC
    """), 
        (my_pos, my_pos)
    )

    pending_actions = c.fetchall()
    recent_leaves = my_leaves[:3]

    conn.close()

    return render_template(
        "user_dashboard.html",
        my_leaves=my_leaves,
        recent_leaves=recent_leaves,
        my_mc=my_mc,
        pending_actions=pending_actions,
        my_pos=my_pos,
        entitlement=entitlement,
        used_leave=used,
        balance_leave=balance
    )

# ================= LEAVE TYPE CANONICAL MAP =================
LEAVE_TYPE_MAP = {
    # ===== Annual Leave =====
    "NORMAL": "AL",
    "EMERGENCY": "AL",

    # ===== Compassionate Leave =====
    "DEATH OF IMMEDIATE FAMILY MEMBERS": "CL",
    "DISASTER (FLOOD/FIRE)": "CL",

    # ===== Others =====
    "LEAVE-IN-LIEU": "LIL",
    "UNPAID LEAVE": "UL",
    "MATERNITY/PATERNITY": "MP",
    "SPECIAL PAID LEAVE": "SPL",

    # ===== Medical =====
    "MC": "MC"
}
@app.route("/user/leave/<int:leave_id>")
@login_required
def user_leave_details(leave_id):
    conn = get_db()
    c = conn.cursor()

    c.execute(
        ("""SELECT l.*, u.full_name, u.position, u.email, u.phone, 
               d.name AS department_name
        FROM leave_applications l
        JOIN users u ON u.id = l.user_id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE l.id=%s AND l.user_id=%s
    """, (leave_id, session["user_id"]))
    )

    leave = c.fetchone()
    conn.close()

    return render_template("user_leave_details.html", 
                           datetime=datetime,
                           leave=leave)


@app.route("/user/upload_mc", methods=["POST"])
@login_required
def user_upload_mc():
    mc_number = request.form.get("mc_number", "").strip()
    mc_start  = request.form.get("mc_start") or None
    mc_end    = request.form.get("mc_end") or None
    file      = request.files.get("mc_file")

    if not file or file.filename == "":
        flash("MC file required.", "warning")
        return redirect(url_for("user_dashboard"))

    filename = secure_filename(file.filename)
    ext = filename.rsplit(".", 1)[-1].lower()
    if ext not in {"pdf", "png", "jpg", "jpeg"}:
        flash("Invalid MC file type.", "danger")
        return redirect(url_for("user_dashboard"))

    ts = datetime.now().strftime("%Y%m%d%H%M%S")
    save_name = f"mc_{session['user_id']}_{ts}_{filename}"
    file.save(os.path.join(LEAVE_UPLOAD_FOLDER, save_name))

    conn = get_db()
    c = conn.cursor()
    c.execute(
        ("""INSERT INTO mc_records
        (user_id, mc_number, start_date, end_date, pdf_path, uploaded_by, created_at)
        VALUES (%s,%s,%s,%s,%s,%s,%s)
    """), (
        session["user_id"],
        mc_number,
        mc_start,
        mc_end,
        save_name,
        session["user_id"],
        datetime.utcnow()
    ))

    c.execute((
        "UPDATE users SET availability='MC' WHERE id=%s"),
        (session["user_id"],)
    )

    conn.commit()
    conn.close()

    flash("Medical Certificate uploaded successfully.", "success")
    return redirect(url_for("user_dashboard"))


def normalize(text):
    if not text:
        return ""
    return " ".join(
        text.upper()
        .replace("&", "AND")
        .replace("-", " ")
        .split()
    )

def get_checker_approver(position, department):
    
    pos = normalize(position)
    dept = normalize(department)

    # ================= SPECIAL POSITIONS (NO CHECKER, ONLY APPROVER) =================
    special_positions = {
        "HEAD OF MANAGEMENT": "GM HR AND ESG",
        "SGM WD": "CEO",
        "GM FIN": "CEO",
        "GM FINANCE": "CEO",
        "GM HR AND ESG": "CEO",
        "HEAD OF CA AND C": "SGM WD",
        "GM ED": "CEO",
        "ENERGY DIVISION": "CEO"
    }

    if pos in special_positions:
        return None, special_positions[pos], "Pending Approval"


    # ================= OTHER SPECIAL CASES =================
    if pos == "HR MANAGER":
        return None, "GM HR AND ESG", "Pending Approval"

    if pos == "DGM SHE":
        return "SGM WD", "GM HR AND ESG", "Pending Recommender"


    # ================= NORMAL STAFF BY DEPARTMENT =================
    dept_rules = {
        "MANAGEMENT": ("GM HR AND ESG", "CEO"),
        "WATER DIVISION": ("SGM WD", "GM HR AND ESG"),
        "FINANCE": ("GM FIN", "GM HR AND ESG"),
        "HR AND ESG": ("HR MANAGER", "GM HR AND ESG"),
        "CORPORATE AFFAIRS AND COMMUNICATIONS": ("HEAD OF CA AND C", "GM HR AND ESG"),
        "CA AND C": ("HEAD OF CA AND C", "GM HR AND ESG"),
        "SAFETY HEALTHY AND ENVIRONMENTALS": ("DGM SHE", "GM HR AND ESG"),
        "SHE": ("DGM SHE", "GM HR AND ESG"),
        "ENERGY DIVISION": ("GM ED", "GM HR AND ESG"),
    }

    if dept in dept_rules:
        checker, approver = dept_rules[dept]
        return checker, approver, "Pending Recommender"

@app.route("/apply_leave", methods=["GET", "POST"])
@login_required
def apply_leave():

    user_id = session.get("user_id")
    full_name = session.get("full_name")
    raw_position = session.get("position")

    pos_upper = normalize(raw_position)

    conn = get_db()
    cur = conn.cursor()

    # get department
    cur.execute(("""SELECT d.name AS department
        FROM users u
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE u.id = %s
    """), (user_id,)
    )
    row = cur.fetchone()
    department = row["department"] if row else ""

    checker_name, approver_name, status = get_checker_approver(pos_upper, department)
    checker_name = normalize(checker_name)
    approver_name = normalize(approver_name)

    if request.method == "POST":

        leave_type = request.form.get("leave_type")
        start_date = request.form.get("start_date")
        end_date = request.form.get("end_date")
        reason = request.form.get("reason")
        contact_address = request.form.get("contact_address")
        contact_phone = request.form.get("contact_phone")

        total_days = calculate_working_days(start_date, end_date)

        support_doc = None
        file = request.files.get("support_doc")
        if file and file.filename:
            filename = secure_filename(file.filename)
            support_doc = filename
            file.save(os.path.join(LEAVE_UPLOAD_FOLDER, support_doc))

        cur.execute(
            ("""INSERT INTO leave_applications
            (user_id, full_name, position, leave_type, start_date, end_date,
            total_days, reason, contact_address, contact_phone, status, checker_name, approver_name, created_at)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """), (
            user_id,
            full_name,
            pos_upper,
            leave_type,
            start_date,
            end_date,
            total_days,
            reason,
            contact_address,
            contact_phone,
            status,
            checker_name,
            approver_name,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ))

        conn.commit()
        conn.close()

        flash("Leave submitted successfully.", "success")
        return redirect(url_for("user_dashboard"))

    # ================= GET REMAINING LEAVE (FIXED) =================

    # get entitlement (FIXED VALUE)
    cur.execute(
        ("SELECT entitlement FROM users WHERE id=%s"), 
        (user_id,)
    )
    user = cur.fetchone()
    entitlement = user["entitlement"] or 0

    # get used leave (SINGLE SOURCE OF TRUTH)
    used_leave = get_used_leave_days(user_id)

    # calculate balance
    remaining_leave = max(0, entitlement - used_leave)

    conn.close()

    return render_template(
        "apply_leave.html",
        full_name=full_name,
        current_date=datetime.now().strftime("%d/%m/%Y"),
        remaining_leave=remaining_leave,
        checker_name=checker_name,
        approver_name=approver_name,
        status=status
    )

@app.route("/approval_dashboard")
@login_required
def approval_dashboard():

    my_position = normalize(session.get("position"))

    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        ("""SELECT 
            la.*,
            u.full_name,
            u.position,
            d.name AS department_name
        FROM leave_applications la
        JOIN users u ON la.user_id = u.id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE 
            (
                la.status = 'Pending Recommender'
                AND UPPER(TRIM(la.checker_name)) = %s
            )
            OR
            (
                la.status = 'Pending Approval'
                AND UPPER(TRIM(la.approver_name)) = %s
            )
        ORDER BY la.created_at DESC
    """), (my_position, my_position)
    )

    leaves = cur.fetchall()
    conn.close()

    return render_template("approval_dashboard.html", leaves=leaves)

@app.route("/leave/<int:leave_id>/check", methods=["POST"])
@login_required
def check_leave_action(leave_id):

    my_pos = normalize(session.get("position"))

    conn = get_db()
    c = conn.cursor()

    c.execute(
        ("SELECT * FROM leave_applications WHERE id=%s"), 
        (leave_id,)
    )
    leave = c.fetchone()

    if not leave:
        flash("Leave not found.", "danger")
        return redirect(url_for("approval_dashboard"))

    if leave["status"] != "Pending Recommender":
        flash("Leave already processed.", "warning")
        return redirect(url_for("approval_dashboard"))

    if normalize(leave["checker_name"]) != my_pos:
        flash("You are not authorized to check this leave.", "danger")
        return redirect(url_for("approval_dashboard"))

    c.execute(
        ("""UPDATE leave_applications
        SET status='Pending Approval',
            checked_at=%s
        WHERE id=%s
    """), (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), leave_id))

    conn.commit()
    conn.close()

    flash("Leave checked and sent to approver.", "success")
    return redirect(url_for("user_dashboard"))  

@app.route("/leave/<int:leave_id>/approve", methods=["POST"])
@login_required
def approve_leave_action(leave_id):

    my_pos = normalize(session.get("position"))
    conn = get_db()
    c = conn.cursor()

    c.execute(
        ("SELECT * FROM leave_applications WHERE id=%s"), 
        (leave_id,)
    )
    leave = c.fetchone()

    if not leave:
        flash("Leave not found.", "danger")
        return redirect(url_for("approval_dashboard"))

    if leave["status"] != "Pending Approval":
        flash("Leave already processed.", "warning")
        return redirect(url_for("approval_dashboard"))

    # ✅ Approve
    c.execute(
        ("""UPDATE leave_applications
        SET status='Approved',
            approved_at=%s
        WHERE id=%s
    """), (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), leave_id))

    conn.commit()
    conn.close()

    flash("Leave approved and balance updated.", "success")
    return redirect(url_for("user_dashboard"))


@app.route("/leave/<int:leave_id>/reject", methods=["POST"])
@login_required
def reject_leave_action(leave_id):

    my_pos = normalize(session.get("position"))

    conn = get_db()
    c = conn.cursor()

    c.execute(
        ("SELECT * FROM leave_applications WHERE id=%s"), 
        (leave_id,)
    )
    leave = c.fetchone()

    if not leave:
        flash("Leave not found.", "danger")
        return redirect(url_for("approval_dashboard"))

    if normalize(leave["checker_name"]) != my_pos and normalize(leave["approver_name"]) != my_pos:
        flash("You are not authorized to reject this leave.", "danger")
        return redirect(url_for("approval_dashboard"))

    c.execute(
        ("""UPDATE leave_applications
        SET status='Rejected',
            approved_at=%s
        WHERE id=%s
    """), (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), leave_id))

    conn.commit()
    conn.close()

    flash("Leave rejected.", "info")
    return redirect(url_for("user_dashboard"))  


@app.route("/debug/positions")
def debug_positions():
    conn = get_db()
    c = conn.cursor()
    c.execute(
        ("SELECT id, full_name, position FROM users"))
    rows = c.fetchall()
    conn.close()

    output = "<h3>Users & Positions in DB</h3><ul>"
    for r in rows:
        output += f"<li>ID: {r['id']}, Name: {r['full_name']}, Position Stored: '{r['position']}'</li>"
    output += "</ul>"
    return output

@app.route("/calendar")
@login_required
def calendar():
    conn = get_db()
    c = conn.cursor()

    # Fetch holidays (still global)
    c.execute(
        ("SELECT name, date FROM holidays ORDER BY date"))
    holidays = c.fetchall()

    # 🔹 Upcoming leaves for this user only
    c.execute(
        ("""SELECT l.id, l.start_date, l.end_date, u.full_name, l.leave_type, l.status
        FROM leaves l
        JOIN users u ON u.id = l.user_id
        WHERE l.user_id = %s
        AND date(l.end_date) >= date('now')
        ORDER BY l.start_date ASC
    """), (session["user_id"],)
    )
    leaves = c.fetchall()

    # 🔹 Leave history (past leaves for this user only)
    c.execute(
        ("""SELECT l.id, l.start_date, l.end_date, u.full_name, l.leave_type, l.status
        FROM leaves l
        JOIN users u ON u.id = l.user_id
        WHERE l.user_id = %s
        AND date(l.end_date) < date('now')
        ORDER BY l.start_date DESC
    """), (session["user_id"],)
    )
    leave_history = c.fetchall()

    conn.close()
    return render_template(
        "calendar.html",
        holidays=holidays,
        leaves=leaves,
        leave_history=leave_history
    )

@app.route("/profile", methods=["GET", "POST"])
@login_required
def profile():
    conn = get_db()
    c = conn.cursor()

    # ================= POST: UPDATE PASSWORD =================
    if request.method == "POST":
        password = request.form.get("password", "").strip()

        if password:
            c.execute(
                ("UPDATE users SET password_hash=%s WHERE id=%s"),
                (generate_password_hash(password), session["user_id"])
            )

        conn.commit()
        conn.close()
        flash("Profile updated successfully.", "success")
        return redirect(url_for("profile"))

    # ================= GET USER DATA =================
    c.execute(
        ("""SELECT u.*, d.name AS department_name
        FROM users u
        LEFT JOIN departments d ON d.id = u.department_id
        WHERE u.id=%s
    """, (session["user_id"],))
    )
    user = c.fetchone()

    entitlement = user["entitlement"] or 0

    # ================= CALCULATE USED LEAVE (FIXED) =================
    used_leave = get_used_leave_days(session["user_id"])

    remaining_leave = max(0, entitlement - used_leave)

    conn.close()

    return render_template(
        "profile.html",
        user=user,
        remaining_leave=remaining_leave,
        hide_balance=user["role"].upper() in ["ADMIN", "CEO"]
                     or (user["position"] or "").upper() == "CEO"
    )


from flask import jsonify, url_for

@app.route("/update_profile_photo", methods=["POST"])
@login_required
def update_profile_photo():
    file = request.files.get("profile_photo")

    if not file:
        return jsonify({"error": "No file"}), 400

    ext = file.filename.rsplit(".",1)[1].lower()
    filename = f"user_{session['user_id']}.{ext}"

    save_path = os.path.join(app.root_path, "static", "uploads", "profile_photos", filename)
    file.save(save_path)

    conn = get_db()
    c = conn.cursor()
    c.execute(
        ("UPDATE users SET profile_photo=%s WHERE id=%s"),
        (filename, session["user_id"])
    )
    conn.commit()
    conn.close()

    image_url = url_for("static", filename="uploads/profile_photos/" + filename)

    print("IMAGE SAVED:", image_url)  # DEBUG

    return jsonify({"url": image_url})


@app.route('/delete_profile_photo', methods=['POST'])
@login_required
def delete_profile_photo():

    user_id = session['user_id']

    conn = get_db()
    c = conn.cursor()

    c.execute(
        ("SELECT profile_photo FROM users WHERE id=%s"), 
        (user_id,)
    )
    row = c.fetchone()

    if row and row["profile_photo"]:
        path = os.path.join(app.config["PROFILE_UPLOAD_FOLDER"], row["profile_photo"])
        if os.path.exists(path):
            os.remove(path)

    c.execute(
        ("UPDATE users SET profile_photo=NULL WHERE id=%s"), 
        (user_id,)
    )
    conn.commit()
    conn.close()

    default_url = "https://ui-avatars.com/api/?name=User&background=random"

    return jsonify({"url": default_url})


@app.route("/settings", methods=["GET", "POST"])
@admin_required
def settings():
    conn = get_db()
    c = conn.cursor()
    if request.method == "POST":
        org_name = request.form.get("org_name", "J-Leave App")
        theme = request.form.get("theme", "blue")
        c.execute(
            ("INSERT INTO settings (key, value) VALUES ('org_name', %s) ON CONFLICT(key) DO UPDATE SET value=excluded.value", (org_name,)))
        c.execute(
            ("INSERT INTO settings (key, value) VALUES ('theme', %s) ON CONFLICT(key) DO UPDATE SET value=excluded.value", (theme,)))
        conn.commit()
        conn.close()
        flash("Settings saved.", "success")
        return redirect(url_for("settings"))
    c.execute(
        ("SELECT key, value FROM settings"))
    settings_rows = {row["key"]: row["value"] for row in c.fetchall()}
    conn.close()
    return render_template("settings.html", settings=settings_rows)

@app.route("/admin/leaves/assign-checker/<int:leave_id>", methods=["POST"])
@admin_required
def assign_checker(leave_id):
    """Admin assigns a specific user to check this leave."""
    checker_id = request.form.get("checker_id")
    if not checker_id:
        flash("Please select a checker.", "warning")
        return redirect(url_for("manage_leaves"))

    conn = get_db()
    c = conn.cursor()
    c.execute(
        ("SELECT full_name, position FROM users WHERE id=%s"), 
        (checker_id,)
    )
    checker = c.fetchone()

    if not checker:
        flash("Invalid checker selected.", "danger")
        conn.close()
        return redirect(url_for("manage_leaves"))

    c.execute(
        ("""UPDATE leaves 
        SET checked_by_user_id=%s, checked_by_position=%s, checked_status='Pending'
        WHERE id=%s
    """), (checker_id, checker["position"], leave_id)
    )
    conn.commit()
    conn.close()

    flash(f"Checker assigned: {checker['full_name']} ({checker['position']}).", "success")
    return redirect(url_for("manage_leaves"))


@app.route("/admin/leaves/assign-approver/<int:leave_id>", methods=["POST"])
@admin_required
def assign_approver_user(leave_id):
    """Admin assigns a specific user to approve this leave."""
    approver_id = request.form.get("approver_id")
    if not approver_id:
        flash("Please select an approver.", "warning")
        return redirect(url_for("manage_leaves"))

    conn = get_db()
    c = conn.cursor()
    c.execute(
        ("SELECT full_name, position FROM users WHERE id=%s"), 
        (approver_id,)
    )
    approver = c.fetchone()

    if not approver:
        flash("Invalid approver selected.", "danger")
        conn.close()
        return redirect(url_for("manage_leaves"))

    c.execute(
        ("""UPDATE leaves 
        SET approved_by_user_id=%s, next_approver=%s, next_approver_position=%s, status='Pending'
        WHERE id=%s
    """), (approver_id, approver["full_name"], approver["position"], leave_id)
    )
    conn.commit()
    conn.close()

    flash(f"Approver assigned: {approver['full_name']} ({approver['position']}).", "success")
    return redirect(url_for("manage_leaves"))

@app.route("/toggle-theme")
def toggle_theme():
    current = session.get("theme_mode", "light")
    session["theme_mode"] = "dark" if current == "light" else "light"
    return redirect(request.referrer or url_for("user_dashboard"))


# --------------- simple API for live clock (optional) ---------------
# @app.route("/api/server-time")
# def server_time():
#     return jsonify({"server_time": datetime.now().isoformat()})
@app.route("/api/leave-trend")
def leave_trend_api():
    view = request.args.get("view", "weekly")
    if view == "monthly":
        labels = ["Jan", "Feb", "Mar", "Apr"]
        annual = [5, 8, 6, 7]
        sick = [2, 3, 1, 4]
    else:
        labels = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        annual = [2, 1, 3, 1, 0, 0, 2]
        sick = [0, 1, 0, 1, 1, 0, 0]
    return {"labels": labels, "annual": annual, "sick": sick}

@app.route("/leave/<int:leave_id>")
@login_required
def leave_details(leave_id):

    conn = get_db()
    c = conn.cursor()

    c.execute(
        ("""SELECT l.*, u.full_name, u.position, d.name AS department_name
        FROM leave_applications l
        JOIN users u ON u.id = l.user_id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE l.id=%s
    """), (leave_id,)
    )
    leave = c.fetchone()
    conn.close()

    my_pos = normalize(session.get("position"))

    can_check = (
        leave["status"] == "Pending Recommender"
        and normalize(leave["checker_name"]) == my_pos
    )

    can_approve = (
        leave["status"] == "Pending Approval"
        and normalize(leave["approver_name"]) == my_pos
    )

    return render_template(
        "leave_details.html",
        leave=leave,
        can_check=can_check,
        can_approve=can_approve,
        datetime=datetime
    )


@app.route("/api/leave/<int:leave_id>")
@login_required
def api_leave_details(leave_id):
    """Return leave details as JSON for modal display."""
    conn = get_db()
    c = conn.cursor()
    c.execute(
        ("""SELECT 
            l.*, 
            u.full_name, u.position, u.email, u.phone, u.address,
            d.name AS department_name
        FROM leaves l
        JOIN users u ON l.user_id = u.id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE l.id=%s
    """), (leave_id,)
    )
    leave = c.fetchone()
    conn.close()

    if not leave:
        return jsonify({"error": "Leave not found"}), 404

    return jsonify(dict(leave))

@app.route("/debug/leaves")
def debug_leaves():
    conn = get_db()
    c = conn.cursor()
    c.execute(
        ("""SELECT id, full_name, position, status, checker_name, approver_name
        FROM leave_applications
        ORDER BY id DESC
    """))
    rows = c.fetchall()
    conn.close()

    out = "<h2>Leave Applications DEBUG</h2><table border=1 cellpadding=5>"
    out += "<tr><th>ID</th><th>Name</th><th>Position</th><th>Status</th><th>Checker</th><th>Approver</th></tr>"
    for r in rows:
        out += f"<tr><td>{r['id']}</td><td>{r['full_name']}</td><td>{r['position']}</td><td>{r['status']}</td><td>{r['checker_name']}</td><td>{r['approver_name']}</td></tr>"
    out += "</table>"
    return out

def get_departments():
    conn = sqlite3.connect("database.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute(
        ("""
        SELECT DISTINCT department
        FROM users
        WHERE department IS NOT NULL AND department != ''
        ORDER BY department
    """))

@app.route("/leave_docs/<path:filename>")
@login_required
def leave_docs(filename):
    # Semua user login boleh view, kalau nak stricter, check role sini
    return send_from_directory(LEAVE_UPLOAD_FOLDER, filename)

import smtplib
from email.mime.text import MIMEText

def send_notification(user_id, message):
    conn = get_db()
    c = conn.cursor()
    c.execute(
        ("SELECT email, phone FROM users WHERE id=%s"), 
        (user_id,)
    )
    user = c.fetchone()
    conn.close()

    if not user:
        return

    if user["email"]:
        send_email(user["email"], "J-Leave Notification", message)

    if user["phone"]:
        send_whatsapp(user["phone"], message)

def send_email(to_email, subject, message):
    sender = "noreply@jleave.com"
    msg = MIMEText(message)
    msg["Subject"] = subject
    msg["From"] = sender
    msg["To"] = to_email

    try:
        smtp = smtplib.SMTP("smtp.gmail.com", 587)
        smtp.starttls()
        smtp.login("YOUR_EMAIL@gmail.com", "YOUR_APP_PASSWORD")
        smtp.sendmail(sender, [to_email], msg.as_string())
        smtp.quit()
    except Exception as e:
        print("Email error:", e)


def send_whatsapp(phone, message):
    print(f"WhatsApp sent to {phone}: {message}")
    # Integrate Twilio / Fonnte later here

@app.route("/leave/file/<filename>")
@login_required
def leave_file(filename):
    return send_from_directory(LEAVE_UPLOAD_FOLDER, filename)

@app.route("/leave/<int:leave_id>/download/pdf")
@login_required
def download_leave_pdf(leave_id):
    conn = get_db()
    c = conn.cursor()
    c.execute(
        ("""SELECT l.*, u.full_name, u.position, u.email, u.phone, u.address,
               d.name AS department_name
        FROM leave_applications l
        JOIN users u ON l.user_id = u.id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE l.id=%s
    """), 
        (leave_id,)
    )
    leave = c.fetchone()
    conn.close()

    if not leave:
        flash("Leave record not found.", "danger")
        return redirect(url_for("user_dashboard"))

    # Render HTML template
    html = render_template("leave_details.html", leave=leave, datetime=datetime.now())

    # Convert HTML → PDF
    pdf = HTML(string=html, base_url=request.host_url).write_pdf()

    return send_file(
        io.BytesIO(pdf),
        as_attachment=True,
        download_name=f"leave_application_{leave_id}.pdf",
        mimetype="application/pdf"
    )

@app.route("/forgot", methods=["GET", "POST"])
def forgot_password():
    if request.method == "POST":
        email = request.form.get("email").strip()

        conn = get_db()
        c = conn.cursor()
        c.execute(
            ("SELECT id FROM users WHERE email=%s"), 
            (email,)
        )
        user = c.fetchone()

        if not user:
            flash("Email not found.", "danger")
            return redirect(url_for("forgot_password"))

        token = str(uuid4())
        expiry = (datetime.utcnow() + timedelta(hours=1)).isoformat()

        c.execute(
            ("""UPDATE users
            SET reset_token=%s, reset_token_expiry=%s
            WHERE email=%s
        """), (token, expiry, email)
        )
        conn.commit()
        conn.close()

        reset_link = url_for("reset_password", token=token, _external=True)

        send_email(
            email, "Reset Your J-Leave Account", f"Click here: {reset_link}" )

        flash("Reset link sent to your email.", "info")
        return redirect(url_for("login"))

    return render_template("forgot_password.html")

@app.route("/reset/<token>", methods=["GET", "POST"])
def reset_password(token):
    conn = get_db()
    c = conn.cursor()

    c.execute(
        ("SELECT id, reset_token_expiry FROM users WHERE reset_token=%s"), 
        (token,)
    )
    user = c.fetchone()

    if not user:
        flash("Invalid or expired link.", "danger")
        return redirect(url_for("login"))

    expiry = datetime.fromisoformat(user["reset_token_expiry"])
    if datetime.utcnow() > expiry:
        flash("Link expired.", "warning")
        return redirect(url_for("forgot_password"))

    if request.method == "POST":
        username = request.form.get("username").strip()
        p1 = request.form.get("password")
        p2 = request.form.get("password2")

        if p1 != p2:
            flash("Passwords do not match.", "danger")
            return redirect(request.url)

        if len(p1) < 6:
            flash("Password must be at least 6 characters.", "danger")
            return redirect(request.url)

        password_hash = generate_password_hash(p1)

        if username:
            c.execute(
                ("""UPDATE users
                SET username=%s, password_hash=%s, reset_token=NULL, reset_token_expiry=NULL
                WHERE id=%s
            """), (username, password_hash, user["id"])
            )
        else:
            c.execute(
                ("""UPDATE users
                SET password_hash=%s, reset_token=NULL, reset_token_expiry=NULL
                WHERE id=%s
            """), (password_hash, user["id"])
            )

        conn.commit()
        conn.close()

        flash("Password updated. You may now login.", "success")
        return redirect(url_for("login"))

    return render_template("reset_password.html")


from flask import send_file
import io

@app.route("/export/leave/excel")
def export_excel():
    mode = request.args.get("mode", "weekly")
    data = get_leave_data(mode)

    wb = Workbook()
    ws = wb.active
    ws.title = "Leave Report"

    ws.append(["Employee Name", "Leave Type", "Start Date", "End Date", "Status"])

    for row in data:
        ws.append(row)

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    filename = f"LeaveReport_{mode}.xlsx"
    return send_file(file_stream, as_attachment=True, download_name=filename)

# Downloadable leave report
# INDIVIDUAL LEAVE REPORT

from datetime import datetime, timedelta
from flask import render_template, request

@app.route("/admin/leave-report/employee/<int:user_id>/view")
@admin_required
def view_individual_leave_report(user_id):

    year = request.args.get("year")
    if not year:
        year = str(datetime.now().year)

    conn = get_db()
    c = conn.cursor()

    # ================= EMPLOYEE =================
    c.execute(
        ("""SELECT u.full_name, u.position, u.entitlement, d.name AS department
        FROM users u
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE u.id = %s
    """), (user_id,)
    )
    emp = c.fetchone()

    # ================= APPROVED LEAVES =================
    c.execute(
        ("""SELECT leave_type, start_date, end_date, total_days
        FROM leave_applications
        WHERE user_id = %s
          AND status = 'Approved'
          AND EXTRACT(YEAR FROM la.start_date)  = %s
        ORDER BY start_date
    """), (user_id, year)
    )
    approved_leaves = c.fetchall()

    # ================= MC RECORDS (REAL DATA) =================
    c.execute(
        ("""SELECT mc_number, start_date, end_date
        FROM mc_records
        WHERE user_id = %s
          AND EXTRACT(YEAR FROM start_date) = %s
        ORDER BY start_date
    """), (user_id, year)
    )
    mc_records = c.fetchall()

    conn.close()

    # ================= MONTHS =================
    months = ["JAN","FEB","MAR","APR","MAY","JUN",
              "JUL","AUG","SEP","OCT","NOV","DEC"]

    month_map = {
        "JAN":"01","FEB":"02","MAR":"03","APR":"04",
        "MAY":"05","JUN":"06","JUL":"07","AUG":"08",
        "SEP":"09","OCT":"10","NOV":"11","DEC":"12"
    }

    # ================= MONTHLY LEAVE (NO MC) =================
    monthly = {m: {} for m in months}

    for l in approved_leaves:
        if not l["start_date"] or not l["end_date"]:
            continue
        if l["leave_type"] == "MC":
            continue

        start = datetime.strptime(l["start_date"], "%Y-%m-%d").date()
        end   = datetime.strptime(l["end_date"], "%Y-%m-%d").date()

        d = start
        while d <= end:
            if d.year == int(year) and d.weekday() < 5:
                m = months[d.month - 1]
                monthly[m][l["leave_type"]] = monthly[m].get(l["leave_type"], 0) + 1
            d += timedelta(days=1)

    # ================= SUMMARY =================
    entitled = emp["entitlement"] or 0
    used = sum(sum(v.values()) for v in monthly.values())
    balance = max(0, entitled - used)

    summary = {
        "entitled": entitled,
        "used": used,
        "balance": balance
    }

    # ================= LEAVE TYPES =================
    leave_types = sorted(
        {l["leave_type"] for l in approved_leaves if l["leave_type"] != "MC"}
    )

    # ================= FORMAT APPROVED LEAVES =================
    formatted_approved = []
    for l in approved_leaves:
        formatted_approved.append({
            "leave_type": l["leave_type"],
            "start_date": l["start_date"],
            "end_date": l["end_date"],
            "total_days": l["total_days"]
        })

    # ================= FORMAT MC (FOLLOW UPLOAD DATA) =================
    formatted_mc = []
    for mc in mc_records:
        formatted_mc.append({
            "mc_number": mc["mc_number"],
            "start_date": mc["start_date"],
            "end_date": mc["end_date"]
        })

    return render_template(
        "reports/individual_leave_report.html",
        employee={
            "name": emp["full_name"],
            "position": emp["position"]
        },
        department=emp["department"],
        year=year,
        printed_date=datetime.now().strftime("%d %b %Y"),
        current_year=datetime.now().year,

        summary=summary,
        months=months,
        month_map=month_map,

        monthly=monthly,
        leave_types=leave_types,

        approved_leaves=formatted_approved,
        mc_records=formatted_mc
    )


from weasyprint import HTML
from datetime import datetime

@app.route("/leave-report/individual/<int:user_id>/pdf")
def download_individual_leave_report_pdf(user_id):
    year = request.args.get("year", datetime.now().year)

    data = build_individual_leave_report(user_id, year)

    html = render_template(
        "reports/individual_leave_report.html",
        **data,
        printed_date=datetime.now().strftime("%d %b %Y")
    )

    pdf = HTML(string=html).write_pdf()

    return Response(
        pdf,
        headers={
            "Content-Type": "application/pdf",
            "Content-Disposition": f"attachment; filename=leave_{user_id}_{year}.pdf"
        }
    )


from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from flask import send_file
import io

@app.route("/leave-report/individual/<int:user_id>/excel")
def download_individual_leave_report_excel(user_id):
    year = request.args.get("year", datetime.now().year)

    data = build_individual_leave_report(user_id, year)

    return generate_individual_leave_excel(data, f"leave_{user_id}_{year}.xlsx")


def build_individual_leave_report(user_id, year):
    conn = get_db()
    cur = conn.cursor()

    # Employee info
    cur.execute(
        ("""SELECT 
            u.full_name,
            u.position,
            d.name AS department
        FROM users u
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE u.id = %s
    """), (user_id,)
    )
    emp = cur.fetchone()

    if not emp:
        conn.close()
        raise ValueError("Employee not found")

    monthly = {m: 0 for m in [
        "January","February","March","April","May","June",
        "July","August","September","October","November","December"
    ]}

    cur.execute(
        ("""SELECT start_date, end_date
        FROM leave_applications
        WHERE user_id = %s
          AND status = 'Approved'
          AND EXTRACT(YEAR FROM start_date) = %s
    """), (user_id, int(year))
    )

    leaves = cur.fetchall()
    conn.close()

    for l in leaves:
        start = datetime.strptime(l["start_date"], "%Y-%m-%d").date()
        end   = datetime.strptime(l["end_date"], "%Y-%m-%d").date()

        for d in daterange(start, end):
            if d.year == int(year) and d.weekday() < 5:
                monthly[d.strftime("%B")] += 1

    used = sum(monthly.values())
    entitled = 14

    return {
        "employee": {
            "name": emp["full_name"],
            "department": emp["department"],
            "position": emp["position"]
        },
        "monthly": monthly,
        "summary": {
            "entitled": entitled,
            "used": used,
            "balance": max(0, entitled - used)
        }
    }

from collections import defaultdict
from datetime import datetime, timedelta

def daterange(start, end):
    for n in range((end - start).days + 1):
        yield start + timedelta(n)

def build_employee_leave_matrix(leaves, year):
    months = ['JAN','FEB','MAR','APR','MAY','JUN','JUL','AUG','SEP','OCT','NOV','DEC']

    from collections import defaultdict
    monthly = {m: defaultdict(int) for m in months}
    monthly_details = {m: defaultdict(list) for m in months}

    total_used = 0

    for l in leaves:
        # ✅ FIX 1
        if l["status"] != "Approved":
            continue

        # ✅ FIX 2
        start = datetime.strptime(l["start_date"], "%Y-%m-%d").date()
        end   = datetime.strptime(l["end_date"], "%Y-%m-%d").date()

        for d in daterange(start, end):
            if d.year != int(year):
                continue

            m = months[d.month - 1]

            # ✅ FIX 3
            monthly[m][l["leave_type"]] += 1

            monthly_details[m][l["leave_type"]].append({
                "date": d.strftime("%d-%m-%Y"),
                "days": 1
            })

            # ✅ FIX 4
            if l["leave_type"] != "MC":
                total_used += 1

    return {
        "monthly": monthly,
        "monthly_details": monthly_details,
        "total_used": total_used
    }
    
from flask import render_template, request, send_file
from datetime import datetime
import pandas as pd
from io import BytesIO
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet

@app.route("/team_leave_report")
@admin_required
def team_leave_report():

    department = request.args.get("department", "").strip()

    conn = get_db()
    c = conn.cursor()

    # ================= DEPARTMENT LIST =================
    c.execute(
        ("SELECT name FROM departments ORDER BY name"))
    departments = [r["name"] for r in c.fetchall()]

    report = []

    if department:
        c.execute(
            ("""SELECT
                u.id AS user_id,
                u.full_name,
                u.position,
                u.entitlement,
                la.leave_type,
                la.start_date,
                la.end_date,
                la.total_days
            FROM leave_applications la
            JOIN users u ON u.id = la.user_id
            LEFT JOIN departments d ON u.department_id = d.id
            WHERE la.status = 'Approved'
              AND d.name = %s
            ORDER BY u.full_name, la.start_date
        """), (department,)
        )

        rows = c.fetchall()

        employees = {}

        for r in rows:
            uid = r["user_id"]

            if uid not in employees:
                employees[uid] = {
                    "name": r["full_name"],
                    "position": r["position"],
                    "entitlement": r["entitlement"] or 0,
                    "used": 0,
                    "leaves": []
                }

            # MC does NOT deduct entitlement
            if r["leave_type"] != "MC":
                employees[uid]["used"] += r["total_days"] or 0

            employees[uid]["leaves"].append({
                "leave_type": r["leave_type"],
                "date": f"{r['start_date']} → {r['end_date']}",
                "days": r["total_days"] or 0
            })

        # Final formatting
        for emp in employees.values():
            report.append({
                "name": emp["name"],
                "position": emp["position"],
                "leaves": emp["leaves"],
                "used": emp["used"],
                "balance": max(0, emp["entitlement"] - emp["used"])
            })

    conn.close()

    return render_template(
        "team_leave_report.html",
        departments=departments,
        selected_department=department,
        report=report
    )
    

from datetime import datetime, timedelta, date
import calendar

def get_leave_report_data(year, month, department):
    first_day = date(year, month, 1)
    last_day = date(year, month, calendar.monthrange(year, month)[1])

    conn = get_db()
    cur = conn.cursor()

    params = [last_day.strftime("%Y-%m-%d"), first_day.strftime("%Y-%m-%d")]

    dept_filter = ""
    if department != "all":
        dept_filter = "AND d.name = %s"
        params.append(department)

    # ================= APPROVED LEAVES =================
    cur.execute(
        (f"""SELECT
            u.id AS user_id,
            u.full_name,
            u.entitlement,
            d.name AS department,
            la.leave_type,
            la.start_date,
            la.end_date
        FROM leave_applications la
        JOIN users u ON u.id = la.user_id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE la.status = 'Approved'
          AND la.start_date <= %s
          AND la.end_date >= %s
          {dept_filter}
        ORDER BY u.full_name, la.start_date
    """), params
    )

    rows = cur.fetchall()
    users = {}

    for r in rows:
        uid = r["user_id"]

        if uid not in users:
            users[uid] = {
                "full_name": r["full_name"],
                "department": r["department"],
                "entitlement": r["entitlement"] or 0,
                "remaining": r["entitlement"] or 0,
                "daily": {}
            }

        lt_raw = (r["leave_type"] or "").strip().upper()
        code = LEAVE_TYPE_MAP.get(lt_raw)

        if not code:
            continue

        start = max(datetime.strptime(r["start_date"], "%Y-%m-%d").date(), first_day)
        end   = min(datetime.strptime(r["end_date"], "%Y-%m-%d").date(), last_day)

        cur_day = start
        while cur_day <= end:
            if cur_day.weekday() < 5:
                day_no = cur_day.day

                if users[uid]["daily"].get(day_no) != "MC":
                    users[uid]["daily"][day_no] = code

                if code != "MC":
                    users[uid]["remaining"] -= 1

            cur_day += timedelta(days=1)

    # ================= MC RECORDS =================
    cur.execute(
        ("""SELECT user_id, start_date, end_date
        FROM mc_records
        WHERE start_date <= %s
          AND end_date >= %s
    """), (
        last_day.strftime("%Y-%m-%d"),
        first_day.strftime("%Y-%m-%d")
    )
    )

    mc_rows = cur.fetchall()

    for m in mc_rows:
        uid = m["user_id"]
        if uid not in users:
            continue

        start = max(datetime.strptime(m["start_date"], "%Y-%m-%d").date(), first_day)
        end   = min(datetime.strptime(m["end_date"], "%Y-%m-%d").date(), last_day)

        cur_day = start
        while cur_day <= end:
            if cur_day.weekday() < 5:
                users[uid]["daily"][cur_day.day] = "MC"
            cur_day += timedelta(days=1)

    conn.close()

    return list(users.values())

@app.route("/leave-report/department/preview")
@admin_required
def preview_leave_report_department():
    from datetime import datetime, timedelta, date
    import calendar

    year = int(request.args.get("year", datetime.now().strftime("%Y")))
    month = int(request.args.get("month", datetime.now().strftime("%m")))
    department = request.args.get("department", "all")

    # ===== Month range =====
    first_day = date(year, month, 1)
    last_day = date(year, month, calendar.monthrange(year, month)[1])
    days_in_month = calendar.monthrange(year, month)[1]

    conn = get_db()
    cur = conn.cursor()

    params = [
        last_day.strftime("%Y-%m-%d"),
        first_day.strftime("%Y-%m-%d")
    ]

    dept_filter = ""
    if department != "all":
        dept_filter = "AND d.name = %s"
        params.append(department)

    # ================= APPROVED LEAVES =================
    cur.execute(
        (f"""SELECT
            u.id AS user_id,
            u.full_name,
            u.entitlement,
            d.name AS department,
            la.leave_type,
            la.start_date,
            la.end_date
        FROM leave_applications la
        JOIN users u ON u.id = la.user_id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE la.status = 'Approved'
          AND la.start_date <= %s
          AND la.end_date>= %s
          {dept_filter}
        ORDER BY u.full_name, la.start_date
    """), params
    )

    rows = cur.fetchall()
    users = {}

    # ================= BUILD USER STRUCTURE =================
    for r in rows:
        uid = r["user_id"]

        if uid not in users:
            users[uid] = {
                "full_name": r["full_name"],
                "department": r["department"],
                "entitlement": r["entitlement"] or 0,
                "remaining": r["entitlement"] or 0,
                "daily": {}   # { day(int): "AL"/"CL"/"MC"/... }
            }

        # Normalize leave type
        lt_raw = (r["leave_type"] or "").strip().upper()
        code = LEAVE_TYPE_MAP.get(lt_raw)

        if not code:
            continue

        start = max(datetime.strptime(r["start_date"], "%Y-%m-%d").date(), first_day)
        end   = min(datetime.strptime(r["end_date"], "%Y-%m-%d").date(), last_day)

        cur_day = start
        while cur_day <= end:
            if cur_day.weekday() < 5:  # working days only
                day_no = cur_day.day

                # MC should not be overwritten
                if users[uid]["daily"].get(day_no) != "MC":
                    users[uid]["daily"][day_no] = code

                # Deduct entitlement except MC
                if code != "MC":
                    users[uid]["remaining"] -= 1

            cur_day += timedelta(days=1)

    # ================= MC RECORDS (APPLIED DATE) =================
    cur.execute(
        ("""SELECT
            m.user_id,
            m.start_date,
            m.end_date
        FROM mc_records m
        WHERE m.start_date <= %s
          AND m.end_date >= %s
    """), (
        last_day.strftime("%Y-%m-%d"),
        first_day.strftime("%Y-%m-%d")
    )
    )

    mc_rows = cur.fetchall()

    for m in mc_rows:
        uid = m["user_id"]

        if uid not in users:
            continue  # only users in this report

        start = max(datetime.strptime(m["start_date"], "%Y-%m-%d").date(), first_day)
        end   = min(datetime.strptime(m["end_date"], "%Y-%m-%d").date(), last_day)

        cur_day = start
        while cur_day <= end:
            if cur_day.weekday() < 5:
                day_no = cur_day.day
                # MC overrides everything
                users[uid]["daily"][day_no] = "MC"

            cur_day += timedelta(days=1)

    conn.close()

    return render_template(
        "reports/leave_report_department_preview.html",
        rows=list(users.values()),
        year=year,
        month=f"{month:02d}",
        department=department,
        printed_date=date.today().strftime("%d-%m-%Y"),
        days_in_month=days_in_month
    )
    
from weasyprint import HTML
from flask import make_response
from datetime import datetime, timedelta, date
import calendar

@app.route("/leave-report/department/pdf")
@admin_required
def team_leave_pdf():

    year = int(request.args.get("year"))
    month = int(request.args.get("month"))
    department = request.args.get("department", "all")

    first_day = date(year, month, 1)
    last_day = date(year, month, calendar.monthrange(year, month)[1])
    days_in_month = calendar.monthrange(year, month)[1]

    conn = get_db()
    cur = conn.cursor()

    params = [last_day.strftime("%Y-%m-%d"), first_day.strftime("%Y-%m-%d")]
    dept_filter = ""
    if department != "all":
        dept_filter = "AND d.name=%s"
        params.append(department)

    cur.execute(
        (f"""SELECT
            u.id AS user_id,
            u.full_name,
            u.entitlement,
            d.name AS department,
            la.leave_type,
            la.start_date,
            la.end_date
        FROM leave_applications la
        JOIN users u ON u.id = la.user_id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE la.status='Approved'
          AND la.start_date <= %s
          AND la.end_date >= %s
          {dept_filter}
        ORDER BY u.full_name, la.start_date
    """), params
    )

    rows = cur.fetchall()
    users = {}

    for r in rows:
        uid = r["user_id"]
        if uid not in users:
            users[uid] = {
                "full_name": r["full_name"],
                "department": r["department"],
                "entitlement": r["entitlement"] or 0,
                "remaining": r["entitlement"] or 0,
                "daily": {}
            }

        code = LEAVE_TYPE_MAP.get((r["leave_type"] or "").upper())
        if not code:
            continue

        start = max(datetime.strptime(r["start_date"], "%Y-%m-%d").date(), first_day)
        end = min(datetime.strptime(r["end_date"], "%Y-%m-%d").date(), last_day)

        cur_day = start
        while cur_day <= end:
            if cur_day.weekday() < 5:
                day_no = cur_day.day
                if users[uid]["daily"].get(day_no) != "MC":
                    users[uid]["daily"][day_no] = code
                if code != "MC":
                    users[uid]["remaining"] -= 1
            cur_day += timedelta(days=1)

    # MC override
    cur.execute(
        ("""SELECT user_id, start_date, end_date
        FROM mc_records
        WHERE start_date <= %s AND end_date >= %s
    """), 
        (last_day.strftime("%Y-%m-%d"), first_day.strftime("%Y-%m-%d"))
    )
    mc_rows = cur.fetchall()

    for m in mc_rows:
        uid = m["user_id"]
        if uid not in users:
            continue
        start = max(datetime.strptime(m["start_date"], "%Y-%m-%d").date(), first_day)
        end = min(datetime.strptime(m["end_date"], "%Y-%m-%d").date(), last_day)

        d = start
        while d <= end:
            if d.weekday() < 5:
                users[uid]["daily"][d.day] = "MC"
            d += timedelta(days=1)

    conn.close()

    html = render_template(
        "reports/leave_report_department_preview.html",
        rows=list(users.values()),
        year=year,
        month=f"{month:02d}",
        department=department,
        printed_date=date.today().strftime("%d-%m-%Y"),
        days_in_month=days_in_month,
        current_year=datetime.now().year,
        pdf_mode=True   # IMPORTANT
    )

    pdf = HTML(string=html, base_url=request.root_url).write_pdf()

    response = make_response(pdf)
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = f"attachment; filename=Employee_Leave_Report_{department}_{month}_{year}.pdf"
    return response

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import calendar, io
from flask import send_file

LEAVE_COLORS = {
    "AL":"4B74FF",
    "CL":"F4A623",
    "LIL":"2ECC71",
    "UL":"B0B0B0",
    "MP":"FF66B2",
    "MC":"FF3B30",
    "SPL":"9B59B6"
}


from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
import calendar, io
from flask import send_file
from datetime import date

LEAVE_COLORS = {
    "AL":"4B74FF",
    "CL":"F4A623",
    "LIL":"2ECC71",
    "UL":"B0B0B0",
    "MP":"FF66B2",
    "MC":"FF3B30",
    "SPL":"9B59B6"
}

@app.route("/leave-report/department/excel")
@admin_required
def team_leave_excel():

    year = int(request.args.get("year"))
    month = int(request.args.get("month"))
    department = request.args.get("department","all")

    users = get_leave_report_data(year, month, department)

    wb = Workbook()
    ws = wb.active
    ws.title = "Leave Report"

    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))

    header_blue = PatternFill("solid", fgColor="4B74FF")
    table_header_fill = PatternFill("solid", fgColor="40B6FF")

    title_font = Font(bold=True, color="FFFFFF", size=12)
    subtitle_font = Font(bold=True, size=11)

    days_in_month = calendar.monthrange(year, month)[1]
    total_cols = 3 + days_in_month + 1  # NAME, DEPT, ENT, days, BALANCE

    # ================= LOGO =================
    try:
        img = ExcelImage("static/logo jetamaa.jpg")
        img.height = 60
        img.width = 120
        ws.add_image(img, "A1")
    except:
        pass

    # ================= TITLE =================
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=total_cols)
    cell = ws.cell(row=1, column=2, value="EMPLOYEE LEAVE REPORT")
    cell.fill = header_blue
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

    MONTH_ABBR = {
        "01":"JAN","02":"FEB","03":"MAR","04":"APR",
        "05":"MAY","06":"JUN","07":"JUL","08":"AUG",
        "09":"SEP","10":"OCT","11":"NOV","12":"DEC"
    }

    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=total_cols)
    sub = ws.cell(row=2, column=2, value=f"{department.upper()}-{MONTH_ABBR[f'{month:02d}']}-{year}")
    sub.font = subtitle_font
    sub.alignment = Alignment(horizontal="center")

    # ================= LEGEND =================
    legend_row = 4
    ws.cell(row=legend_row, column=total_cols-3, value="LEAVE TYPE").font = Font(bold=True)

    legend_items = [
        ("Annual","AL"), ("Unpaid","UL"),
        ("Compassionate","CL"), ("Maternity / Paternity","MP"),
        ("Leave-In-Lieu","LIL"), ("Special Paid Leave","SPL"),
        ("Medical (MC)","MC")
    ]

    r = legend_row + 1
    for name, code in legend_items:
        color_cell = ws.cell(row=r, column=total_cols-3)
        color_cell.fill = PatternFill("solid", fgColor=LEAVE_COLORS[code])
        color_cell.border = thin

        ws.cell(row=r, column=total_cols-2, value=name)
        r += 1

    # ================= TABLE HEADER =================
    start_table_row = r + 2

    headers = ["NAME","DEPARTMENT","ENTITLEMENT"] + \
              [str(i) for i in range(1, days_in_month+1)] + ["BALANCE"]

    ws.append(headers)

    for col in range(1, len(headers)+1):
        c = ws.cell(row=start_table_row, column=col)
        c.fill = table_header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = thin

    # ================= TABLE DATA =================
    row_no = start_table_row + 1

    for u in users:
        ws.append([u["full_name"], u["department"], u["entitlement"]])

        for d in range(1, days_in_month+1):
            cell = ws.cell(row=row_no, column=3+d)
            code = u["daily"].get(d)

            if code:
                cell.fill = PatternFill("solid", fgColor=LEAVE_COLORS.get(code,"FFFFFF"))

            cell.border = thin
            cell.alignment = Alignment(horizontal="center")

        ws.cell(row=row_no, column=3+days_in_month+1, value=u["remaining"])
        row_no += 1

    # ================= FOOTER =================
    footer_row = row_no + 2
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=total_cols)
    footer = ws.cell(row=footer_row, column=1,
        value=f"Printed on {date.today().strftime('%d-%m-%Y')} · Jetama Sdn Bhd © {year}")
    footer.alignment = Alignment(horizontal="right")

    # ================= COLUMN WIDTH =================
    for col in range(1, total_cols+1):
        ws.column_dimensions[get_column_letter(col)].width = 4

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 18

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name=f"Employee_Leave_Report_{department}_{month:02d}_{year}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.context_processor
def inject_current_year():
    from datetime import datetime
    return {
        "current_year": datetime.now().year
    }

@app.context_processor
def inject_my_leaves():
    my_leaves = []
    if session.get("user_id"):
        conn = get_db()
        c = conn.cursor()
        c.execute(
            ("""SELECT status
            FROM leave_applications
            WHERE user_id=%s
            ORDER BY id DESC
            LIMIT 1
        """), 
        (session["user_id"],)
        )
        
        my_leaves = c.fetchall()
        conn.close()
    return dict(my_leaves=my_leaves)

@app.context_processor
def inject_globals():
    return dict(current_time=datetime.now().strftime("%Y-%m-%d %H:%M"))

@app.route("/mark_notifications_seen", methods=["POST"])
@login_required
def mark_notifications_seen():
    session["notif_seen"] = True
    return "", 204


@app.route("/print_monthly_matrix_pdf")
def print_monthly_matrix_pdf():
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from datetime import datetime, timedelta
    import io

    matrix_month = request.args.get("matrix_month")
    matrix_year = request.args.get("matrix_year")
    department = request.args.get("matrix_department", "all")

    # ================= VALIDATION =================
    if not matrix_month or not matrix_year:
        return "Error: matrix_month and matrix_year are required parameters", 400

    matrix_month = str(matrix_month).zfill(2)

    try:
        first_day = datetime.strptime(f"{matrix_year}-{matrix_month}-01", "%Y-%m-%d")
    except ValueError:
        return "Invalid date format for matrix_month or matrix_year", 400

    # ================= MONTH NAME =================
    MONTH_NAMES = {
        "01":"January","02":"February","03":"March","04":"April",
        "05":"May","06":"June","07":"July","08":"August",
        "09":"September","10":"October","11":"November","12":"December"
    }

    month_name = MONTH_NAMES.get(matrix_month, matrix_month)

    # ================= LAST DAY =================
    if int(matrix_month) == 12:
        last_day = datetime(int(matrix_year) + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = datetime(int(matrix_year), int(matrix_month) + 1, 1) - timedelta(days=1)

    conn = get_db()
    cur = conn.cursor()

    users = {}

    # ================= APPROVED LEAVES =================
    params = [last_day.strftime("%Y-%m-%d"), first_day.strftime("%Y-%m-%d")]
    dept_filter = ""

    if department != "all":
        dept_filter = "AND d.name = %s"
        params.append(department)

    cur.execute(
        (f"""SELECT
            la.user_id,
            u.full_name,
            la.leave_type,
            la.start_date,
            la.end_date
        FROM leave_applications la
        JOIN users u ON u.id = la.user_id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE la.status='Approved'
          AND la.start_date <= %s
          AND la.end_date >= %s
          {dept_filter}
        ORDER BY u.full_name
    """), params
    )

    for r in cur.fetchall():
        if not r["start_date"] or not r["end_date"]:
            continue

        uid = r["user_id"]
        users.setdefault(uid, {"user_name": r["full_name"], "leaves": {}})

        start = datetime.strptime(r["start_date"], "%Y-%m-%d")
        end   = datetime.strptime(r["end_date"], "%Y-%m-%d")

        cur_day = max(start, first_day)
        while cur_day <= min(end, last_day):
            users[uid]["leaves"][cur_day.strftime("%d")] = r["leave_type"]
            cur_day += timedelta(days=1)

    # ================= MC RECORDS =================
    params = [last_day.strftime("%Y-%m-%d"), first_day.strftime("%Y-%m-%d")]
    dept_filter = ""

    if department != "all":
        dept_filter = "AND d.name = %s"
        params.append(department)

    cur.execute(
        (f"""SELECT
            m.user_id,
            u.full_name,
            m.start_date,
            m.end_date
        FROM mc_records m
        JOIN users u ON u.id = m.user_id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE m.start_date<= %s
          AND m.end_date >= %s
          {dept_filter}
    """), params
    )

    for m in cur.fetchall():
        if not m["start_date"] or not m["end_date"]:
            continue

        uid = m["user_id"]
        users.setdefault(uid, {"user_name": m["full_name"], "leaves": {}})

        start = datetime.strptime(m["start_date"], "%Y-%m-%d")
        end   = datetime.strptime(m["end_date"], "%Y-%m-%d")

        cur_day = max(start, first_day)
        while cur_day <= min(end, last_day):
            users[uid]["leaves"][cur_day.strftime("%d")] = "MC"
            cur_day += timedelta(days=1)

    conn.close()

    # ================= BUILD PDF =================
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=20,
        leftMargin=20,
        topMargin=30,
        bottomMargin=20
    )

    styles = getSampleStyleSheet()
    elements = []

    title = f"Monthly On-Leave List - {month_name} {matrix_year} (Department: {department})"
    elements.append(Paragraph(f"<b>{title}</b>", styles["Title"]))
    elements.append(Spacer(1, 12))

    header = ["Name"] + [f"{i:02d}" for i in range(1, 32)]
    data = [header]

    table_style = TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.lightblue),
        ("GRID", (0,0), (-1,-1), 0.5, colors.black),
        ("ALIGN", (1,1), (-1,-1), "CENTER"),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
    ])

    for user in users.values():
        row = [Paragraph(user["user_name"], styles["Normal"])]

        for d in range(1, 32):
            day = f"{d:02d}"
            if day in user["leaves"]:
                if user["leaves"][day] == "MC":
                    row.append(Paragraph("<font color='red'>✓</font>", styles["Normal"]))
                else:
                    row.append(Paragraph("<font color='green'>✓</font>", styles["Normal"]))
            else:
                row.append("")
        data.append(row)

    table = Table(data, colWidths=[150] + [18]*31, repeatRows=1)
    table.setStyle(table_style)

    elements.append(table)
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(
        "<b>Legend:</b> <font color='green'>✓ Approved Leave</font> | <font color='red'>✓ MC Leave</font>",
        styles["Normal"]
    ))

    doc.build(elements)
    buffer.seek(0)

    filename = f"Monthly_On_Leave_{month_name}_{matrix_year}_{department}.pdf"

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf"
    )

if __name__ == "__main__":
    init_db()
    app.run(debug=True, host="0.0.0.0", port=5000)
    
